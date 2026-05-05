"""
detect_stockouts.py — Stockout detection from POS SOH report
Foodland Wudinna

Usage:
    python detect_stockouts.py <path_to_SOH_file.xlsx>
    — or —
    Drag & drop the SOH file onto "Launch Stockout Detector.bat"

Reads a SOH export, identifies items that ran out of stock, estimates
lost trading days and revenue, and appends results to:
    05_waste/Stockout_Log.csv

De-duplicates by (report_date, item_name) — re-running on the same
SOH file overwrites those rows with the latest calculation.
"""

import re
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT         = Path(__file__).parent
CALENDAR     = ROOT / "07_powerbi/data/dim_calendar.csv"
STOCKOUT_LOG = ROOT / "05_waste/Stockout_Log.csv"

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# ── Config ─────────────────────────────────────────────────────────────────────
SOH_MIN_THRESHOLD = -500   # exclude bulk/kg POS data errors (e.g. -187,952)
MAX_LAST_SOLD_AGE = 30     # item must have sold within this many days to qualify
LOOKBACK_WEEKS    = 8      # sales history window for daily qty / price calc


# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s)).strip()


def trading_days_between(start: pd.Timestamp, end: pd.Timestamp,
                         cal: pd.DataFrame) -> int:
    """
    Count store-open trading days strictly between start (exclusive)
    and end (exclusive).  Uses is_store_open from dim_calendar.
    """
    if pd.isna(start) or pd.isna(end) or start >= end:
        return 0
    mask = (
        (cal["date"] > start) &
        (cal["date"] < end) &
        (cal["is_store_open"] == 1)
    )
    return int(mask.sum())


def next_delivery_after(report_date: pd.Timestamp,
                        cal: pd.DataFrame) -> pd.Timestamp:
    """
    Return the first delivery day (Tuesday or Friday) after report_date
    that is also a store-open day in dim_calendar.
    Delivery schedule: Tuesday AM and Friday AM.
    """
    open_days = cal[
        (cal["is_store_open"] == 1) &
        (cal["date"] > report_date)
    ].sort_values("date")

    delivery_days = open_days[open_days["date"].dt.dayofweek.isin([1, 4])]

    if not delivery_days.empty:
        return delivery_days.iloc[0]["date"]

    # Fallback: next Tuesday if calendar data runs out
    days_ahead = (1 - report_date.dayofweek) % 7 or 7
    return report_date + pd.Timedelta(days=days_ahead)


# ── SOH parser ────────────────────────────────────────────────────────────────
def parse_soh_report(path: Path) -> pd.DataFrame:
    """
    Parse the POS SOH Excel export.

    Expected columns (row 6 header):
      [0] GTIN  [3] Description  [7] AWS  [8] Stock On Hand
      [14] Net Retail  [17] Last Sold
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Locate header row
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "GTIN":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError(f"Could not find GTIN header row in {path.name}.")

    skip_prefixes = {"department", "total", "ezi-manager"}
    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        desc = row[3]
        if desc is None:
            continue
        desc_str = str(desc).strip().lower()
        if any(desc_str.startswith(p) for p in skip_prefixes):
            continue
        if not isinstance(row[8], (int, float)):
            continue

        records.append({
            "name":       norm(desc),
            "aws":        row[7],
            "soh":        row[8],
            "net_retail": row[14],
            "last_sold":  row[17],
        })

    df = pd.DataFrame(records)
    df["aws"]        = pd.to_numeric(df["aws"],        errors="coerce").fillna(0)
    df["soh"]        = pd.to_numeric(df["soh"],        errors="coerce")
    df["net_retail"] = pd.to_numeric(df["net_retail"], errors="coerce").fillna(0)
    df["last_sold"]  = pd.to_datetime(df["last_sold"], format="%d %b %y",
                                      errors="coerce")
    return df


def infer_report_date(path: Path) -> pd.Timestamp:
    """
    Infer SOH report date from filename (expects SOH_DDMMYYYY.xlsx).
    Falls back to today if pattern not found.
    """
    match = re.search(r"(\d{8})", path.stem)
    if match:
        try:
            return pd.to_datetime(match.group(1), format="%d%m%Y")
        except ValueError:
            pass
    print(f"  Warning: could not parse date from filename '{path.name}'. Using today.")
    return pd.Timestamp.today().normalize()


# ── Sales data ────────────────────────────────────────────────────────────────
def load_sales() -> pd.DataFrame:
    """Load sales from SQLite via db.py, returning columns: Date, Name, Qty, Rev."""
    from db import load_sales as _db_load
    df = _db_load()
    if df.empty:
        return pd.DataFrame(columns=["Date", "Name", "Qty", "Rev"])
    df = df.rename(columns={"Revenue": "Rev"})
    return df[["Date", "Name", "Qty", "Rev"]]


def build_sales_rates(sales: pd.DataFrame) -> pd.DataFrame:
    """
    Compute per-item avg daily qty and avg unit price from recent sales.
    Window: last LOOKBACK_WEEKS weeks.
    Returns DataFrame with columns: name, daily_qty, avg_price.
    """
    if sales.empty:
        return pd.DataFrame(columns=["name", "daily_qty", "avg_price"])

    cutoff = sales["Date"].max() - pd.Timedelta(weeks=LOOKBACK_WEEKS)
    recent = sales[sales["Date"] >= cutoff].copy()
    n_days = max((sales["Date"].max() - cutoff).days, 1)

    agg = (
        recent.groupby("Name")
        .agg(total_qty=("Qty", "sum"), total_rev=("Rev", "sum"))
        .reset_index()
    )
    agg["daily_qty"] = (agg["total_qty"] / n_days).round(3)
    agg["avg_price"] = np.where(
        agg["total_qty"] > 0,
        (agg["total_rev"] / agg["total_qty"]).round(2),
        0,
    )
    return agg.rename(columns={"Name": "name"})[["name", "daily_qty", "avg_price"]]


# ── Main detection ─────────────────────────────────────────────────────────────
def detect_stockouts(soh_path: Path) -> pd.DataFrame:
    """
    Full pipeline: parse SOH → filter → calculate lost days/revenue.
    Returns a tidy DataFrame ready for the log.
    """
    soh         = parse_soh_report(soh_path)
    report_date = infer_report_date(soh_path)
    today       = pd.Timestamp.today().normalize()

    print(f"  Report date : {report_date.strftime('%d %b %Y')}")
    print(f"  Items in SOH: {len(soh)}")

    # ── Filter genuine stockouts ──────────────────────────────────────────────
    cutoff_last_sold = today - pd.Timedelta(days=MAX_LAST_SOLD_AGE)

    genuine = soh[
        (soh["soh"] <= 0) &
        (soh["soh"] > SOH_MIN_THRESHOLD) &   # exclude POS bulk-kg errors
        (soh["aws"] > 0) &                    # active item (has recent sales)
        (soh["last_sold"].notna()) &
        (soh["last_sold"] >= cutoff_last_sold) &
        (soh["last_sold"] < report_date)      # confirmed ran out before report
    ].copy()

    if genuine.empty:
        print("  No genuine stockouts found.")
        return pd.DataFrame()

    print(f"  Genuine stockouts: {len(genuine)}")

    # ── Supporting data ───────────────────────────────────────────────────────
    cal           = pd.read_csv(CALENDAR)
    cal["date"]   = pd.to_datetime(cal["date_key"].astype(str), format="%Y%m%d")

    sales         = load_sales()
    rates         = build_sales_rates(sales)
    next_delivery = next_delivery_after(report_date, cal)

    print(f"  Next delivery: {next_delivery.strftime('%d %b %Y')}")

    # ── Merge sales rates; fall back to SOH AWS + Net Retail ─────────────────
    # Normalise names to uppercase for matching (same as sales CSV)
    rates["name_key"] = rates["name"].str.upper()
    genuine["name_key"] = genuine["name"].str.upper()

    rates_key = rates[["name_key", "daily_qty", "avg_price"]].copy()
    genuine = genuine.merge(rates_key, on="name_key", how="left")
    genuine["daily_qty"] = genuine["daily_qty"].fillna(genuine["aws"] / 7)
    genuine["avg_price"] = genuine["avg_price"].fillna(genuine["net_retail"])

    # ── Calculate lost trading days and revenue ───────────────────────────────
    genuine["next_delivery"] = next_delivery
    genuine["lost_days"] = genuine["last_sold"].apply(
        lambda ls: trading_days_between(ls, next_delivery, cal)
    )
    genuine["lost_revenue"] = (
        genuine["lost_days"] * genuine["daily_qty"] * genuine["avg_price"]
    ).round(2)

    # ── Format output ─────────────────────────────────────────────────────────
    result = pd.DataFrame({
        "report_date":   report_date.date(),
        "item_name":     genuine["name"].values,
        "soh":           genuine["soh"].round(1).values,
        "last_sold":     genuine["last_sold"].dt.date.values,
        "next_delivery": next_delivery.date(),
        "lost_days":     genuine["lost_days"].values,
        "daily_qty":     genuine["daily_qty"].round(2).values,
        "avg_price":     genuine["avg_price"].round(2).values,
        "lost_revenue":  genuine["lost_revenue"].values,
    })

    return result.sort_values("lost_revenue", ascending=False).reset_index(drop=True)


# ── Log writer ────────────────────────────────────────────────────────────────
def save_to_log(new_rows: pd.DataFrame) -> None:
    """
    Append stockout events to STOCKOUT_LOG.
    De-duplicates: existing rows with the same (report_date, item_name)
    are replaced — re-running on the same SOH file is safe.
    """
    if new_rows.empty:
        return

    if STOCKOUT_LOG.exists():
        existing = pd.read_csv(STOCKOUT_LOG)
        report_dates = set(new_rows["report_date"].astype(str))
        item_names   = set(new_rows["item_name"])
        drop_mask = (
            existing["report_date"].astype(str).isin(report_dates) &
            existing["item_name"].isin(item_names)
        )
        combined = pd.concat([existing[~drop_mask], new_rows], ignore_index=True)
    else:
        STOCKOUT_LOG.parent.mkdir(parents=True, exist_ok=True)
        combined = new_rows

    combined.to_csv(STOCKOUT_LOG, index=False)
    print(f"\n  Saved {len(new_rows)} event(s) → {STOCKOUT_LOG.relative_to(ROOT)}")
    print(f"  Total log entries: {len(combined)}")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:  python detect_stockouts.py <path_to_SOH_file.xlsx>")
        print("        Drag & drop the SOH file onto 'Launch Stockout Detector.bat'")
        sys.exit(1)

    soh_path = Path(sys.argv[1])
    if not soh_path.exists():
        print(f"Error: file not found — {soh_path}")
        sys.exit(1)

    print(f"\nStockout Detector — Foodland Wudinna")
    print(f"File: {soh_path.name}")
    print("-" * 50)

    events = detect_stockouts(soh_path)

    if events.empty:
        print("\nNo stockout events to record.")
    else:
        print("\n  Item                                     | Last Sold  | Days Out | Est. Loss")
        print("  " + "-" * 74)
        for _, r in events.iterrows():
            print(f"  {str(r['item_name'])[:40]:40} | {str(r['last_sold']):10} |"
                  f"    {r['lost_days']:2}    | ${r['lost_revenue']:>8.2f}")

        total = events["lost_revenue"].sum()
        print(f"\n  Total estimated lost revenue: ${total:,.2f}")

        save_to_log(events)

    print("\nDone. Press any key to close.")
    input()
