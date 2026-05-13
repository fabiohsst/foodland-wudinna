"""
generate_order_headless.py — Headless Order Sheet Generator
Foodland Wudinna

Runs the full Freshlink order sheet pipeline without Streamlit.
Designed for GitHub Actions — reads sales from a CSV snapshot, not the DB.

Usage:
    python generate_order_headless.py --soh PATH --specials PATH [--output PATH] [--order-type TYPE]

Arguments:
    --soh        Path to the Stock on Hand Excel/CSV file
    --specials   Path to the Freshlink specials bulletin (.docx)
    --output     Output path for the Excel file (default: auto-named in 04_ordering/)
    --order-type Force a specific cycle: WED_FRI or FRI_TUE (default: auto from weekday)

Exit codes:
    0  Success
    1  Fatal error (missing files, bad inputs, etc.)
"""

import argparse
import io
import re
import sys
import zipfile
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

ROOT = Path(__file__).parent

# ── Paths ──────────────────────────────────────────────────────────────────────
SALES_SNAPSHOT   = ROOT / "03_model/sales_snapshot.csv"
CALENDAR_CSV     = ROOT / "07_powerbi/data/dim_calendar.csv"
MODEL_PKL        = ROOT / "03_model/demand_model.pkl"
SPECIALS_MAP_CSV = ROOT / "01_data/reference/specials_mapping.csv"
INPUTS_DIR       = ROOT / "03_model/inputs"
OUTPUT_DIR       = ROOT / "04_ordering"
_API_KEY_FILE    = ROOT / ".api_key"

# ── Constants ──────────────────────────────────────────────────────────────────
LOOKBACK_WEEKS        = 8
EWMA_ALPHA            = 0.35
SOH_MIN               = 0
SOH_MAX               = 300
ACTIVE_LOOKBACK_WEEKS = 2
_ORDER_DAY_WEIGHT     = 0.4
_LLM_API_URL          = "https://api.anthropic.com/v1/messages"
_LLM_MODEL            = "claude-haiku-4-5-20251001"

SUPPLIER_PREFIXES: dict[str, list[str] | None] = {
    "Freshlink":     None,
    "Bowlsome":      ["BOWLSOME", "COMM CO"],
    "Local Kitchen": ["L/C", "L/K", "LK "],
    "Simply Tasty":  ["S/TASTY"],
}
_ALL_NAMED_PREFIXES = [
    p.upper()
    for s, prefs in SUPPLIER_PREFIXES.items() if prefs
    for p in prefs
]

CONSOLIDATION_GROUPS = {
    "Cabbage": {
        "display_name": "CABBAGE (order whole heads)",
        "unit": "whole",
        "variants": {
            "Cabbage Whole":   1.00,
            "Cabbage Half":    0.50,
            "Cabbage Quarter": 0.25,
        },
    },
    "Cabbage Red": {
        "display_name": "CABBAGE RED (order whole heads)",
        "unit": "whole",
        "variants": {
            "Cabbage Red Quarter": 0.25,
        },
    },
    "Cabbage Chinese": {
        "display_name": "CABBAGE CHINESE (order whole heads)",
        "unit": "whole",
        "variants": {
            "Cabbage Chinese Whole": 1.00,
            "Cabbage Chinese Half":  0.50,
        },
    },
    "Cauliflower": {
        "display_name": "CAULIFLOWER (order whole heads)",
        "unit": "whole",
        "variants": {
            "Cauliflower Per Each": 1.00,
            "Cauliflower Half":     0.50,
        },
    },
    "Celery": {
        "display_name": "CELERY (order whole stalks)",
        "unit": "whole",
        "variants": {
            "Celery Large": 1.00,
            "Celery Half":  0.50,
        },
    },
    "Watermelon": {
        "display_name": "WATERMELON (order kg)",
        "unit": "kg",
        "variants": {
            "Watermelon Seedless Per Kg": 1.00,
            "Sliced Watermelon":          1.00,
        },
    },
    "Rockmelon": {
        "display_name": "ROCKMELON (order kg)",
        "unit": "kg",
        "variants": {
            "Rockmelon Per Kg":  1.00,
            "Sliced Rockmelon":  1.00,
        },
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def item_belongs_to(item_name: str, supplier: str) -> bool:
    name_up = norm(item_name).upper()
    prefs   = SUPPLIER_PREFIXES[supplier]
    if prefs is None:
        return not any(name_up.startswith(p) for p in _ALL_NAMED_PREFIXES)
    return any(name_up.startswith(p.upper()) for p in prefs)


def _load_api_key() -> str | None:
    import os
    env = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if env:
        return env
    if _API_KEY_FILE.exists():
        key = _API_KEY_FILE.read_text().strip()
        if key:
            return key
    return None


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load_sales_from_snapshot() -> pd.DataFrame:
    """Load sales history from the CSV snapshot (no DB required)."""
    if not SALES_SNAPSHOT.exists():
        raise FileNotFoundError(
            f"Sales snapshot not found: {SALES_SNAPSHOT}\n"
            "Run export_sales_snapshot.py locally to generate it."
        )
    df = pd.read_csv(SALES_SNAPSHOT, parse_dates=["Date"])
    df["Quantity"] = df["Qty"]
    return df


def load_calendar() -> pd.DataFrame:
    if not CALENDAR_CSV.exists():
        raise FileNotFoundError(f"Calendar CSV not found: {CALENDAR_CSV}")
    cal = pd.read_csv(CALENDAR_CSV)
    cal["date"] = pd.to_datetime(cal["date_key"].astype(str), format="%Y%m%d")
    return cal


def load_model():
    """Load the LightGBM model. Returns None if not found."""
    if not MODEL_PKL.exists():
        return None
    try:
        sys.path.insert(0, str(ROOT))
        from predict import load_model as _load
        return _load()
    except Exception as e:
        print(f"[warn] Could not load model: {e} — falling back to EWMA", file=sys.stderr)
        return None


def load_specials_mapping() -> pd.DataFrame:
    if not SPECIALS_MAP_CSV.exists():
        return pd.DataFrame(columns=["bulletin_description", "pos_name", "supplier", "verified", "notes"])
    df = pd.read_csv(SPECIALS_MAP_CSV)
    df.columns = df.columns.str.strip()
    return df


# ══════════════════════════════════════════════════════════════════════════════
# SOH PARSING
# ══════════════════════════════════════════════════════════════════════════════

def parse_soh_file(file_path: Path) -> dict:
    """
    Parse a SOH Excel/CSV file → {item_name: stock_qty}.
    Raises ValueError with a clear message if the file cannot be parsed.
    """
    import openpyxl

    path = Path(file_path)
    fname = path.name.lower()

    if fname.endswith(".csv"):
        soh = pd.read_csv(path)
    elif fname.endswith((".xlsx", ".xls")):
        raw_bytes = path.read_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        ws = wb.active
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if "gtin" in cells or "description" in cells:
                header_row = i
                break
        if header_row is None:
            header_row = 0
        soh = pd.read_excel(io.BytesIO(raw_bytes), header=header_row)
    else:
        raise ValueError(f"Unsupported SOH file type: {path.suffix}")

    soh.columns = [" ".join(str(c).split()) if c is not None else "" for c in soh.columns]

    _name_candidates  = ("name", "item", "item name", "description", "product", "product name")
    _stock_candidates = ("stock", "qty", "quantity", "on hand", "onhand", "soh",
                         "current stock", "stock on hand", "available")

    name_col  = next((c for c in soh.columns if c.lower() in _name_candidates), None)
    stock_col = next((c for c in soh.columns if c.lower() in _stock_candidates), None)

    if name_col is None:
        raise ValueError(f"No item-name column found in SOH file. Columns: {list(soh.columns)}")
    if stock_col is None:
        raise ValueError(f"No stock column found in SOH file. Columns: {list(soh.columns)}")

    soh = soh.rename(columns={name_col: "Name", stock_col: "Stock"})
    soh["Name"]  = soh["Name"].apply(norm)
    soh["Stock"] = pd.to_numeric(soh["Stock"], errors="coerce").fillna(0)

    mask = soh["Stock"].between(SOH_MIN, SOH_MAX)
    return dict(zip(soh.loc[mask, "Name"], soh.loc[mask, "Stock"].round(0).astype(int)))


# ══════════════════════════════════════════════════════════════════════════════
# SPECIALS BULLETIN PARSING
# ══════════════════════════════════════════════════════════════════════════════

def _clean_bulletin_description(raw: str) -> str:
    s = raw
    s = re.sub(r"\*+[^*]+\*+", " ", s)
    s = re.sub(r"\bAPN\s*[\d\s]+", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b\d+(\.\d+)?\s*(g|kg|ml|l)\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bPer\s+Carton\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b\d+\s*Per\s+Carton\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bNew\s+(Line|Season)\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bPre[-\s]?orders\s+required\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"[–—]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _extract_bulletin_descriptions(doc) -> list[str]:
    _skip_patterns = re.compile(
        r"contact\s+phone|phone\s+number|\d{2}\s\d{4}\s\d{4}"
        r"|^j\s+l\s+king|metcash\s+item\s+no|^rainbow\s+fresh",
        re.IGNORECASE,
    )
    seen, results = set(), []
    for table in doc.tables:
        for row in table.rows:
            cells = row.cells
            if len(cells) < 2:
                continue
            raw = re.sub(r"\s+", " ", cells[1].text).strip()
            if not raw:
                continue
            lower = raw.lower()
            if lower in ("description", "size count", "desc", "item"):
                continue
            if re.match(r"^\$?[\d,.]+\s*(per\s+kg|ea|pack|each)?$", lower):
                continue
            if _skip_patterns.search(lower):
                continue
            if len(raw) < 4:
                continue
            clean = _clean_bulletin_description(raw)
            if len(clean) < 4:
                continue
            if clean not in seen:
                seen.add(clean)
                results.append(clean)
    return results


def parse_specials_bulletin_file(file_path: Path) -> tuple[list[str], str | None]:
    """Parse a specials bulletin .docx file → (descriptions, error_message)."""
    from docx import Document as _DocxDoc
    path = Path(file_path)
    try:
        doc = _DocxDoc(str(path))
        return _extract_bulletin_descriptions(doc), None
    except Exception as e:
        return [], f"Could not read specials bulletin: {e}"


def _llm_match_bulletin(raw_text: str, db_products: list[str], api_key: str) -> tuple[list[str], list[str], str | None]:
    import json as _json, time as _time, requests as _req

    _LLM_SYSTEM = (
        "You are a produce specialist helping an Australian independent supermarket "
        "(Foodland Wudinna, SA) identify which items from a Freshlink weekly specials "
        "bulletin correspond to products in their POS system.\n\n"
        "Always respond with valid JSON only. No commentary, no markdown fences."
    )
    _LLM_USER_TEMPLATE = """\
Below is the text of a Freshlink weekly specials bulletin, followed by the store's
complete Fruit & Veg POS product list.

For every item line in the bulletin (including all sub-variants of multi-variant items),
find the single best matching POS product name.

BULLETIN TEXT:
{bulletin_text}

AVAILABLE POS PRODUCT NAMES (must match verbatim or return null):
{pos_names_block}

Return this exact JSON structure:
{{
  "matched": [
    {{"bulletin_desc": "<description as seen in bulletin>", "pos_name": "<exact POS name from list>"}}
  ],
  "unmatched": ["<bulletin description that has no match>", ...]
}}

Rules:
- pos_name must be copied VERBATIM from the POS list. Do NOT paraphrase or invent names.
- A multi-variant item must produce one matched entry per variant, each with its own pos_name.
- Put into "unmatched" any bulletin item with no reasonable POS equivalent.
- Ignore rows that are headers, supplier contacts, or raw APN numbers.
"""
    payload = {
        "model": _LLM_MODEL,
        "max_tokens": 4096,
        "system": _LLM_SYSTEM,
        "messages": [{"role": "user", "content": _LLM_USER_TEMPLATE.format(
            bulletin_text=raw_text[:8000],
            pos_names_block="\n".join(db_products),
        )}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    last_err = None
    for attempt in range(1, 4):
        try:
            resp = _req.post(_LLM_API_URL, json=payload, headers=headers, timeout=60, verify=False)
            if resp.status_code == 200:
                break
            last_err = f"HTTP {resp.status_code}"
            if resp.status_code in (429, 529):
                _time.sleep(10 * attempt)
                continue
            return [], [], last_err
        except Exception as e:
            last_err = str(e)
            if attempt < 3:
                _time.sleep(5)
    else:
        return [], [], f"API call failed: {last_err}"

    raw = resp.json()["content"][0]["text"].strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    try:
        result = _json.loads(raw)
    except _json.JSONDecodeError as e:
        return [], [], f"LLM returned invalid JSON: {e}"

    db_upper = {n.upper(): n for n in db_products}
    matched, seen = [], set()
    for entry in result.get("matched", []):
        pos = (entry.get("pos_name") or "").strip()
        resolved = db_upper.get(pos.upper())
        if resolved and resolved not in seen:
            seen.add(resolved)
            matched.append(resolved)

    unmatched = [str(u) for u in result.get("unmatched", [])]
    return matched, unmatched, None


def match_bulletin_to_pos(descriptions: list[str], mapping_df: pd.DataFrame,
                           available_names: list[str]) -> tuple[list[str], list[str]]:
    import difflib

    if mapping_df.empty:
        return [], descriptions

    available_upper = {n.upper(): n for n in available_names}
    _STOP = {
        "the", "and", "or", "of", "a", "an", "in", "on", "at", "to", "per",
        "bag", "pack", "punnet", "each", "kg", "g", "ml", "l", "size", "count",
        "punnets", "packs", "new", "season", "line", "fresh",
    }

    def _tokens(s: str) -> set[str]:
        return {w for w in re.sub(r"[^a-z0-9\s]", " ", s.lower()).split()
                if w not in _STOP and len(w) > 2}

    mdf = mapping_df.copy()
    mdf["_key"]    = mdf["bulletin_description"].str.upper().str.strip()
    mdf["_key"]    = mdf["_key"].apply(lambda s: re.sub(r"\s+", " ", s))
    mdf["_tokens"] = mdf["_key"].apply(lambda k: _tokens(k))

    matched, unmatched, seen_pos = [], [], set()

    def _add(pos_name: str) -> None:
        canon    = pos_name.strip()
        resolved = available_upper.get(canon.upper(), canon)
        if resolved not in seen_pos and resolved.upper() in available_upper:
            seen_pos.add(resolved)
            matched.append(resolved)

    for desc in descriptions:
        norm_key  = re.sub(r"\s+", " ", desc.upper().strip())
        desc_toks = _tokens(desc)

        exact = mdf[mdf["_key"] == norm_key]
        if not exact.empty:
            _add(exact.iloc[0]["pos_name"])
            continue

        best_overlap, best_row = 0, None
        for _, mrow in mdf.iterrows():
            mk_toks = mrow["_tokens"]
            if len(mk_toks) == 0:
                continue
            overlap = len(mk_toks & desc_toks)
            if overlap >= 2 and overlap / len(mk_toks) >= 0.5:
                if overlap > best_overlap:
                    best_overlap = overlap
                    best_row = mrow
        if best_row is not None:
            _add(best_row["pos_name"])
            continue

        keys  = mdf["_key"].tolist()
        close = difflib.get_close_matches(norm_key, keys, n=1, cutoff=0.62)
        if close:
            row = mdf[mdf["_key"] == close[0]].iloc[0]
            _add(row["pos_name"])
            continue

        unmatched.append(desc)

    return matched, unmatched


def resolve_specials(specials_path: Path, item_names: list[str]) -> list[str]:
    """
    Full specials resolution pipeline:
    1. Try LLM matching (requires ANTHROPIC_API_KEY)
    2. Fall back to specials_mapping.csv CSV matching
    Returns list of matched POS item names.
    """
    api_key = _load_api_key()
    specials_mapping = load_specials_mapping()

    if api_key:
        # LLM path — extract full text from docx and let the model match
        path = Path(specials_path)
        try:
            raw_bytes = path.read_bytes()
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as z:
                xml = z.read("word/document.xml").decode("utf-8", errors="replace")
            xml  = re.sub(r"<w:p[ />]",  "\n",  xml)
            xml  = re.sub(r"<w:tc[ />]", " | ", xml)
            text = re.sub(r"<[^>]+>",    "",    xml)
            text = re.sub(r"[ \t]+",     " ",   text)
            text = re.sub(r"\n{3,}",     "\n\n", text).strip()
        except Exception as e:
            print(f"[warn] Could not extract text from specials file: {e}", file=sys.stderr)
            text = ""

        if text.strip():
            matched, unmatched, err = _llm_match_bulletin(text, item_names, api_key)
            if err:
                print(f"[warn] LLM matching failed: {err} — falling back to CSV mapping", file=sys.stderr)
            else:
                print(f"[info] Specials matched via LLM: {len(matched)} items, {len(unmatched)} unmatched")
                return matched

    # CSV fallback
    descs, err = parse_specials_bulletin_file(specials_path)
    if err:
        print(f"[warn] Could not parse specials bulletin: {err}", file=sys.stderr)
        return []

    matched, unmatched = match_bulletin_to_pos(descs, specials_mapping, item_names)
    print(f"[info] Specials matched via CSV mapping: {len(matched)} items, {len(unmatched)} unmatched")
    return matched


# ══════════════════════════════════════════════════════════════════════════════
# FORECAST ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def ewma_forecast(sales: pd.DataFrame, items: list, cycle_dates: list, alpha: float = EWMA_ALPHA):
    sales = sales.copy()
    sales["dow"] = sales["Date"].dt.dayofweek
    cycle_labels = [d.strftime("%a %d/%m") for d in cycle_dates]

    forecast = pd.DataFrame({"Name": items})
    for tgt_date, label in zip(cycle_dates, cycle_labels):
        dow = tgt_date.dayofweek
        day_f = []
        for item in items:
            item_df  = sales[sales["Name"] == item]
            same_dow = item_df[item_df["dow"] == dow].sort_values("Date")
            qtys     = same_dow["Quantity"].values
            if len(qtys) == 0:
                day_f.append(round(float(item_df["Quantity"].mean()) if len(item_df) else 0.0, 1))
                continue
            lags    = qtys[-6:][::-1]
            weights = np.array([alpha ** i for i in range(len(lags))])
            weights /= weights.sum()
            day_f.append(round(float(np.dot(weights, lags)), 1))
        forecast[label] = day_f

    forecast["Total Forecast"] = forecast[cycle_labels].sum(axis=1).round(1)
    return forecast, cycle_labels


def compute_forecast(all_sales: pd.DataFrame, active_items: list, cycle_dates: list,
                     specials: list, model_data) -> tuple:
    cycle_labels  = [d.strftime("%a %d/%m") for d in cycle_dates]
    active_cutoff = all_sales["Date"].max() - pd.Timedelta(weeks=ACTIVE_LOOKBACK_WEEKS)
    recently_sold = set(all_sales[all_sales["Date"] >= active_cutoff]["Name"].unique())
    active_filter = recently_sold | set(specials)
    filtered_items = [i for i in active_items if i in active_filter]

    if model_data is not None:
        model_items = set(model_data["active_items"])
        lgbm_items  = [i for i in filtered_items if i in model_items]
        ewma_items  = [i for i in filtered_items if i not in model_items]
    else:
        lgbm_items = []
        ewma_items = filtered_items

    parts = []

    if lgbm_items:
        sys.path.insert(0, str(ROOT))
        from predict import predict_cycle
        lgbm_df, _ = predict_cycle(
            cycle_dates, specials, model_data, all_sales,
            active_filter=active_filter,
        )
        lgbm_df = lgbm_df[lgbm_df["Name"].isin(lgbm_items)].copy()
        parts.append(lgbm_df)

    if ewma_items:
        cutoff      = all_sales["Date"].max() - pd.Timedelta(weeks=LOOKBACK_WEEKS)
        recent_ewma = all_sales[all_sales["Date"] >= cutoff]
        ewma_df, _  = ewma_forecast(recent_ewma, ewma_items, cycle_dates)
        parts.append(ewma_df)

    if parts:
        forecast_df = pd.concat(parts, ignore_index=True)
    else:
        forecast_df = pd.DataFrame({"Name": filtered_items, "Total Forecast": 0.0})
        for label in cycle_labels:
            forecast_df[label] = 0.0

    return forecast_df, cycle_labels, len(lgbm_items), len(ewma_items)


def consolidate_forecast(forecast_df: pd.DataFrame, cycle_labels: list,
                         soh_map: dict) -> pd.DataFrame:
    df = forecast_df.copy()
    rows_to_add = []
    names_to_drop = set()

    for group_key, group in CONSOLIDATION_GROUPS.items():
        display   = group["display_name"]
        unit      = group["unit"]
        variants  = group["variants"]

        norm_variants = {norm(k).lower(): v for k, v in variants.items()}
        matched = df[df["Name"].apply(lambda x: norm(x).lower()).isin(norm_variants.keys())]

        if matched.empty:
            continue

        agg_day = {}
        for label in cycle_labels:
            total = 0.0
            for _, row in matched.iterrows():
                factor = norm_variants.get(norm(row["Name"]).lower(), 1.0)
                total += float(row.get(label, 0) or 0) * factor
            agg_day[label] = round(total, 2)

        raw_total = sum(agg_day.values())

        if unit == "whole":
            order_total = float(np.ceil(raw_total))
        else:
            order_total = round(raw_total, 1)

        consolidated_soh = 0.0
        for variant_name, factor in norm_variants.items():
            for soh_name, soh_qty in soh_map.items():
                if norm(soh_name).lower() == variant_name:
                    consolidated_soh += soh_qty * factor
                    break
        consolidated_soh = round(consolidated_soh, 1) if consolidated_soh > 0 else np.nan

        subdept = matched.iloc[0].get("SubDept", "Other")

        consolidated_row = {
            "Name":           display,
            "Total Forecast": order_total,
            "SubDept":        subdept,
            "Revenue":        matched["Revenue"].sum() if "Revenue" in matched.columns else 0,
            "Sys_Stock":      consolidated_soh if not np.isnan(consolidated_soh) else np.nan,
            "is_system":      not np.isnan(consolidated_soh) if consolidated_soh else False,
            "_consolidated":  True,
            "_unit":          unit,
            "_raw_total":     raw_total,
        }
        for label in cycle_labels:
            consolidated_row[label] = agg_day[label]

        rows_to_add.append(consolidated_row)
        names_to_drop.update(matched["Name"].tolist())

    df = df[~df["Name"].isin(names_to_drop)].copy()
    df["_consolidated"] = False
    df["_unit"]         = "unit"
    df["_raw_total"]    = df["Total Forecast"]

    if rows_to_add:
        df = pd.concat([df, pd.DataFrame(rows_to_add)], ignore_index=True)

    return df


def get_cycle_dates(order_type: str, cal_df: pd.DataFrame, today=None):
    if today is None:
        today = pd.Timestamp.today().normalize()

    cal = cal_df.copy()
    cal["date"] = pd.to_datetime(cal["date"])

    def _next_dow(base, target_dow):
        days = (target_dow - base.dayofweek) % 7 or 7
        return base + pd.Timedelta(days=days)

    def _open_between(start, end):
        mask = (cal["date"] >= start) & (cal["date"] <= end) & (cal["is_store_open"] == 1)
        return sorted(cal.loc[mask, "date"].tolist())

    if order_type == "WED_FRI":
        delivery  = _next_dow(today, 4)
        cycle_end = _next_dow(delivery, 0)
        return _open_between(delivery, cycle_end), delivery

    elif order_type == "FRI_TUE":
        delivery  = _next_dow(today, 1)
        cycle_end = _next_dow(delivery, 3)
        return _open_between(delivery, cycle_end), delivery

    else:
        raise ValueError(f"Unknown order_type: {order_type!r}")


# ══════════════════════════════════════════════════════════════════════════════
# ORDER QTY CALCULATION
# ══════════════════════════════════════════════════════════════════════════════

def calculate_order_quantities(sheet_df: pd.DataFrame, all_sales: pd.DataFrame,
                                cycle_dates: list, delivery_date: pd.Timestamp,
                                cal_df: pd.DataFrame, specials: list,
                                model_data, c_labels: list) -> pd.DataFrame:
    """
    Compute Pre_Del_Fc, Proj_Stock, Order_Qty, Net_Forecast, Order_Qty_Buf.
    Mirrors the inline calculation in app.py.
    """
    today = pd.Timestamp.today().normalize()
    _cal  = cal_df.copy()
    _cal["date"] = pd.to_datetime(_cal["date"])

    _pre_mask = (
        (_cal["date"] >= today) &
        (_cal["date"] <  delivery_date) &
        (_cal["is_store_open"] == 1)
    )
    _pre_delivery_dates = sorted(_cal.loc[_pre_mask, "date"].tolist())

    depletion_map = {}
    if _pre_delivery_dates:
        active_items = sheet_df["Name"].tolist()
        _pre_fc, _pre_labels, _, _ = compute_forecast(
            all_sales, active_items, _pre_delivery_dates, specials, model_data
        )
        _day_weights = {}
        for d, lbl in zip(_pre_delivery_dates, _pre_labels):
            _day_weights[lbl] = _ORDER_DAY_WEIGHT if d == today else 1.0

        _dep_total = pd.Series(0.0, index=_pre_fc.index)
        for lbl, w in _day_weights.items():
            if lbl in _pre_fc.columns:
                _dep_total += _pre_fc[lbl].fillna(0) * w
        _pre_fc["_depletion"] = _dep_total
        depletion_map = dict(zip(_pre_fc["Name"], _pre_fc["_depletion"]))

    sheet_df = sheet_df.copy()
    sheet_df["Pre_Del_Fc"] = sheet_df["Name"].map(depletion_map).fillna(0.0).round(1)
    sheet_df["Proj_Stock"] = sheet_df.apply(
        lambda r: np.nan if pd.isna(r.get("Sys_Stock"))
        else max(0.0, float(r["Sys_Stock"]) - depletion_map.get(r["Name"], 0.0)),
        axis=1,
    )

    def calc_order(row):
        if row.get("_consolidated") and not pd.isna(row.get("Proj_Stock")):
            proj = float(row["Proj_Stock"])
            if row.get("_unit") == "whole":
                return max(0, int(np.ceil(row["_raw_total"])) - int(np.floor(proj)))
            else:
                return max(0, round(row["_raw_total"] - proj, 1))
        elif row.get("_consolidated"):
            return np.nan
        elif not row["is_system"]:
            return np.nan
        else:
            proj = float(row["Proj_Stock"]) if not pd.isna(row.get("Proj_Stock")) else 0
            return max(0, round(row["Total Forecast"] - proj))

    def calc_net_forecast(row):
        if pd.isna(row.get("Sys_Stock")):
            return np.nan
        soh     = float(row["Sys_Stock"])
        pre_del = float(row.get("Pre_Del_Fc", 0) or 0)
        raw     = float(row.get("_raw_total", row["Total Forecast"]))
        return round(raw + pre_del - soh, 1)

    sheet_df["Order_Qty"]    = sheet_df.apply(calc_order, axis=1)
    sheet_df["Net_Forecast"] = sheet_df.apply(calc_net_forecast, axis=1)

    _n_cycle_days = max(len(cycle_dates), 1)

    def calc_buffer(row):
        raw = float(row.get("_raw_total", row["Total Forecast"]))
        buf = raw / _n_cycle_days
        if row.get("_unit") == "whole":
            return max(1, int(np.ceil(buf)))
        return round(buf, 1)

    sheet_df["_buffer"] = sheet_df.apply(calc_buffer, axis=1)
    sheet_df["Order_Qty_Buf"] = sheet_df.apply(
        lambda r: np.nan if pd.isna(r["Order_Qty"])
        else (int(r["Order_Qty"]) + int(r["_buffer"]) if r.get("_unit") == "whole"
              else round(float(r["Order_Qty"]) + float(r["_buffer"]), 1)),
        axis=1,
    )

    return sheet_df


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER (identical logic to app.py build_excel)
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(sheet_df: pd.DataFrame, cycle_labels: list, cycle_dates: list,
                order_type: str) -> bytes:
    n_days = len(cycle_labels)

    HEADER_BG   = "1A5276";  BLUE_ITEM   = "D6EAF8"
    GREY_ITEM   = "F2F3F4";  WHITE_ITEM  = "FFFFFF"
    FORECAST_BG = "EBF5FB";  TOTAL_BG    = "D5E8D4"
    ORDER_BG    = "FCF3CF";  CONSOL_BG   = "E8DAEF"

    def fill(h):
        return PatternFill("solid", fgColor=h)

    thin = Border(
        left=Side("thin", color="CCCCCC"),   right=Side("thin", color="CCCCCC"),
        top=Side("thin", color="CCCCCC"),    bottom=Side("thin", color="CCCCCC"),
    )
    center  = Alignment(horizontal="center", vertical="center")
    left_al = Alignment(horizontal="left",   vertical="center")

    wb = Workbook()

    # ── Tab 1: Order Sheet ─────────────────────────────────────────────────────
    wp = wb.active
    wp.title = "Order Sheet"

    wp.column_dimensions["A"].width = 4
    wp.column_dimensions["B"].width = 36
    wp.column_dimensions["C"].width = 7
    wp.column_dimensions["D"].width = 14
    wp.column_dimensions["E"].width = 13
    wp.column_dimensions["F"].width = 13

    cycle_str_p = " · ".join(d.strftime("%a %d %b") for d in cycle_dates)

    wp.row_dimensions[1].height = 24
    wp.merge_cells("A1:F1")
    c = wp["A1"]
    c.value     = "FRUIT & VEG ORDER — Foodland Wudinna"
    c.font      = Font("Arial", 14, bold=True, color="FFFFFF")
    c.fill      = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    wp.row_dimensions[2].height = 15
    wp.merge_cells("A2:F2")
    wp["A2"].value = (
        f"Delivery: {cycle_dates[0].strftime('%a %d %b %Y') if cycle_dates else '—'}  |  "
        f"Cycle covers: {cycle_str_p}  |  Order: {order_type}"
    )
    wp["A2"].font      = Font("Arial", 9, color="444444")
    wp["A2"].fill      = fill("EAF2FF")
    wp["A2"].alignment = Alignment(horizontal="center", vertical="center")

    wp.row_dimensions[3].height = 13
    wp.merge_cells("A3:F3")
    wp["A3"].value = (
        f"Order Qty includes +1 day safety buffer  "
        f"({n_days}-day cycle → buffer ≈ 1/{n_days} of cycle forecast per item)"
    )
    wp["A3"].font      = Font("Arial", 8, italic=True, color="555555")
    wp["A3"].fill      = fill("FDFEFE")
    wp["A3"].alignment = Alignment(horizontal="center", vertical="center")

    wp.row_dimensions[4].height = 28
    p_hdr_font = Font("Arial", 10, bold=True, color="FFFFFF")
    for ci, hdr in enumerate(["#", "Item Name", "Special", "Cycle\nForecast", "SOH\non Hand", "Order Qty\n(+1 day)"], 1):
        c = wp.cell(4, ci, hdr)
        c.font      = p_hdr_font
        c.fill      = fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin

    p_row = 5;  p_item = 0;  p_subdept = None

    for _, ir in sheet_df.iterrows():
        subdept   = ir["SubDept"]
        is_system = bool(ir["is_system"])
        is_consol = bool(ir.get("_consolidated", False))
        unit_lbl  = ir.get("_unit", "unit")

        if subdept != p_subdept:
            p_subdept = subdept
            wp.row_dimensions[p_row].height = 14
            wp.merge_cells(f"A{p_row}:F{p_row}")
            c = wp.cell(p_row, 1, subdept.upper())
            c.font      = Font("Arial", 9, bold=True, color="FFFFFF")
            c.fill      = fill(HEADER_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for ci in range(1, 7):
                wp.cell(p_row, ci).border = thin
            p_row += 1

        p_item += 1
        row_bg   = fill(CONSOL_BG) if is_consol else (fill(BLUE_ITEM) if is_system else
                   fill(GREY_ITEM) if p_row % 2 == 0 else fill(WHITE_ITEM))
        name_fnt = Font("Arial", 9, bold=(is_system or is_consol),
                        color="6C3483" if is_consol else "000000")

        wp.row_dimensions[p_row].height = 16

        c = wp.cell(p_row, 1, p_item)
        c.font = Font("Arial", 8, color="888888");  c.fill = row_bg
        c.border = thin;  c.alignment = center

        c = wp.cell(p_row, 2, ir["Name"])
        c.font = name_fnt;  c.fill = row_bg
        c.border = thin;  c.alignment = left_al

        special_val = ir.get("Special", "")
        c = wp.cell(p_row, 3, special_val)
        if special_val:
            c.font = Font("Arial", 8, bold=True, color="7D6608")
            c.fill = fill("FEF9E7")
        else:
            c.font = Font("Arial", 8, color="CCCCCC")
            c.fill = row_bg
        c.border = thin;  c.alignment = center

        fc_val  = float(ir.get("_raw_total", ir["Total Forecast"]))
        fc_disp = int(np.ceil(fc_val)) if unit_lbl == "whole" else round(fc_val, 1)
        c = wp.cell(p_row, 4, fc_disp)
        c.font = Font("Arial", 9);  c.fill = fill(FORECAST_BG)
        c.border = thin;  c.alignment = center
        c.number_format = "0" if unit_lbl == "whole" else "0.0"

        if is_system or is_consol:
            soh_val = int(ir["Sys_Stock"]) if not pd.isna(ir.get("Sys_Stock")) else 0
            c = wp.cell(p_row, 5, soh_val)
            c.font = Font("Arial", 9, bold=True,
                          color="6C3483" if is_consol else "1A5276")
        else:
            c = wp.cell(p_row, 5, "")
            c.font = Font("Arial", 9)
        c.fill = row_bg;  c.border = thin;  c.alignment = center

        buf_qty  = ir.get("Order_Qty_Buf")
        has_buf  = not pd.isna(buf_qty) if buf_qty is not None else False
        if has_buf:
            buf_disp = int(buf_qty) if unit_lbl == "whole" else round(float(buf_qty), 1)
            c = wp.cell(p_row, 6, buf_disp)
            c.font = Font("Arial", 10, bold=True,
                          color="C0392B" if buf_disp > 0 else "27AE60")
            c.number_format = "0" if unit_lbl == "whole" else "0.0"
        else:
            c = wp.cell(p_row, 6, "")
            c.font = Font("Arial", 9)
        c.fill = fill(ORDER_BG);  c.border = thin;  c.alignment = center

        p_row += 1

    wp.print_title_rows = "1:4"
    wp.page_setup.orientation = "portrait"
    wp.page_setup.paperSize   = 9
    wp.page_setup.fitToPage   = True
    wp.page_setup.fitToWidth  = 1
    wp.page_setup.fitToHeight = 0
    wp.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    wp.freeze_panes = "D5"

    # ── Tab 2: Detail ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Detail")

    COL_NO     = 1;  COL_NAME   = 2;  COL_SPECIAL = 3
    COL_DAYS   = list(range(4, 4 + n_days))
    COL_SOH    = 4 + n_days;  COL_PRE_FC = 5 + n_days
    COL_PROJ   = 6 + n_days;  COL_TOTAL  = 7 + n_days
    COL_ORDER  = 8 + n_days;  COL_NOTES  = 9 + n_days
    TOTAL_COLS = COL_NOTES

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 7
    for ci in COL_DAYS:
        ws.column_dimensions[get_column_letter(ci)].width = 8
    ws.column_dimensions[get_column_letter(COL_SOH)].width    = 8
    ws.column_dimensions[get_column_letter(COL_PRE_FC)].width = 8
    ws.column_dimensions[get_column_letter(COL_PROJ)].width   = 8
    ws.column_dimensions[get_column_letter(COL_TOTAL)].width  = 9
    ws.column_dimensions[get_column_letter(COL_ORDER)].width  = 9
    ws.column_dimensions[get_column_letter(COL_NOTES)].width  = 22

    ws.row_dimensions[1].height = 24
    ws.merge_cells(f"A1:{get_column_letter(TOTAL_COLS)}1")
    c = ws["A1"]
    c.value = "STOCK COUNT SHEET — Foodland Wudinna | Fruit & Veg"
    c.font  = Font("Arial", 14, bold=True, color="FFFFFF")
    c.fill  = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 16
    ws.merge_cells(f"A2:{get_column_letter(TOTAL_COLS)}2")
    cycle_str = " | ".join(d.strftime("%a %d %b") for d in cycle_dates)
    ws["A2"].value = (
        f"Order date: {date.today().strftime('%d %b %Y')}  |  "
        f"Cycle: {cycle_str}  |  {order_type}"
    )
    ws["A2"].font      = Font("Arial", 9, color="444444")
    ws["A2"].fill      = fill("EAF2FF")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 14
    ws.merge_cells(f"A3:{get_column_letter(TOTAL_COLS)}3")
    ws["A3"].value = (
        "Blue = system-tracked (SOH & Order Qty pre-filled)   "
        "White/Grey = count manually   "
        "Purple = consolidated multi-cut item (total in base unit)"
    )
    ws["A3"].font      = Font("Arial", 8, italic=True, color="555555")
    ws["A3"].fill      = fill("FDFEFE")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[4].height = 32
    hdr_font = Font("Arial", 9, bold=True, color="FFFFFF")
    headers  = (["#", "Item Name", "Special"] + cycle_labels +
                ["SOH\nNow", "Fc to\nDeliv.", "SOH\nat Del.", "Total\nForecast",
                 "Order Qty\n(+1 day)", "Notes"])
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(4, ci, hdr)
        c.font      = hdr_font
        c.fill      = fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin

    row_num = 5;  item_num = 0;  current_subdept = None

    for _, ir in sheet_df.iterrows():
        subdept   = ir["SubDept"]
        is_system = bool(ir["is_system"])
        is_consol = bool(ir.get("_consolidated", False))
        unit_lbl  = ir.get("_unit", "unit")

        if subdept != current_subdept:
            current_subdept = subdept
            ws.row_dimensions[row_num].height = 16
            ws.merge_cells(f"A{row_num}:{get_column_letter(TOTAL_COLS)}{row_num}")
            c = ws.cell(row_num, 1, subdept.upper())
            c.font      = Font("Arial", 10, bold=True, color="FFFFFF")
            c.fill      = fill(HEADER_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for ci in range(1, TOTAL_COLS + 1):
                ws.cell(row_num, ci).border = thin
            row_num += 1

        item_num += 1

        if is_consol:
            row_bg   = fill(CONSOL_BG);  fore_bg   = fill(CONSOL_BG)
            name_fnt = Font("Arial", 9, bold=True, color="6C3483")
        elif is_system:
            row_bg   = fill(BLUE_ITEM);  fore_bg   = fill(BLUE_ITEM)
            name_fnt = Font("Arial", 9, bold=True)
        else:
            row_bg   = fill(GREY_ITEM) if row_num % 2 == 0 else fill(WHITE_ITEM)
            fore_bg  = fill(FORECAST_BG)
            name_fnt = Font("Arial", 9)

        ws.row_dimensions[row_num].height = 16 if not is_consol else 18

        c = ws.cell(row_num, COL_NO, item_num)
        c.font = Font("Arial", 8, color="888888")
        c.fill = row_bg;  c.border = thin;  c.alignment = center

        c = ws.cell(row_num, COL_NAME, ir["Name"])
        c.font = name_fnt
        c.fill = row_bg;  c.border = thin;  c.alignment = left_al

        special_val = ir.get("Special", "")
        c = ws.cell(row_num, COL_SPECIAL, special_val)
        if special_val:
            c.font = Font("Arial", 8, bold=True, color="7D6608")
            c.fill = fill("FEF9E7")
        else:
            c.font = Font("Arial", 8, color="CCCCCC");  c.fill = row_bg
        c.border = thin;  c.alignment = center

        for ci, label in zip(COL_DAYS, cycle_labels):
            val = float(ir.get(label, 0) or 0)
            if unit_lbl == "whole":
                disp = round(val, 2)
            else:
                disp = round(val, 1)
            c = ws.cell(row_num, ci, disp)
            c.font = Font("Arial", 9);  c.fill = fore_bg
            c.border = thin;  c.alignment = center
            c.number_format = "0.00" if unit_lbl == "whole" else "0.0"

        if is_system or is_consol:
            soh_val = int(ir["Sys_Stock"]) if not pd.isna(ir.get("Sys_Stock")) else 0
            c = ws.cell(row_num, COL_SOH, soh_val)
            c.font = Font("Arial", 9, bold=True, color="6C3483" if is_consol else "1A5276")
        else:
            c = ws.cell(row_num, COL_SOH, "");  c.font = Font("Arial", 9)
        c.fill = row_bg;  c.border = thin;  c.alignment = center

        pre_fc_val = ir.get("Pre_Del_Fc")
        c = ws.cell(row_num, COL_PRE_FC,
                    round(float(pre_fc_val), 1) if not pd.isna(pre_fc_val) else "")
        c.font = Font("Arial", 9);  c.fill = row_bg
        c.border = thin;  c.alignment = center

        proj_val = ir.get("Proj_Stock")
        c = ws.cell(row_num, COL_PROJ,
                    round(float(proj_val), 1) if not pd.isna(proj_val) else "")
        c.font = Font("Arial", 9);  c.fill = row_bg
        c.border = thin;  c.alignment = center

        fc_val  = float(ir.get("_raw_total", ir["Total Forecast"]))
        fc_disp = int(np.ceil(fc_val)) if unit_lbl == "whole" else round(fc_val, 1)
        c = ws.cell(row_num, COL_TOTAL, fc_disp)
        c.font = Font("Arial", 9, bold=True);  c.fill = fill(TOTAL_BG)
        c.border = thin;  c.alignment = center
        c.number_format = "0" if unit_lbl == "whole" else "0.0"

        buf_qty = ir.get("Order_Qty_Buf")
        has_buf = not pd.isna(buf_qty) if buf_qty is not None else False
        if has_buf:
            buf_disp = int(buf_qty) if unit_lbl == "whole" else round(float(buf_qty), 1)
            c = ws.cell(row_num, COL_ORDER, buf_disp)
            c.font = Font("Arial", 10, bold=True,
                          color="C0392B" if buf_disp > 0 else "27AE60")
            c.number_format = "0" if unit_lbl == "whole" else "0.0"
        else:
            c = ws.cell(row_num, COL_ORDER, "");  c.font = Font("Arial", 9)
        c.fill = fill(ORDER_BG);  c.border = thin;  c.alignment = center

        c = ws.cell(row_num, COL_NOTES, "")
        c.font = Font("Arial", 9);  c.fill = row_bg
        c.border = thin;  c.alignment = left_al

        row_num += 1

    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.freeze_panes = "D5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def _specials_cycle_start(ref: date) -> date:
    """Return the Wednesday that started the specials cycle containing ref."""
    from datetime import timedelta
    days_since_wed = (ref.weekday() - 2) % 7
    return ref - timedelta(days=days_since_wed)


def find_specials_file(cycle_start: date) -> Path | None:
    """
    Look for specials_YYYY-MM-DD.docx or .doc in INPUTS_DIR for the given cycle start date.
    Returns None if not found.
    """
    stem = f"specials_{cycle_start.strftime('%Y-%m-%d')}"
    for ext in (".docx", ".doc"):
        path = INPUTS_DIR / f"{stem}{ext}"
        if path.exists():
            return path
    return None


def run(soh_path: Path | None, specials_path: Path | None, output_path: Path | None,
        order_type: str | None, specials_next_path: Path | None = None):
    today = pd.Timestamp.today()
    print(f"[info] Starting order generation — {today.strftime('%A %d %b %Y %H:%M')}")

    # ── Default SOH path ───────────────────────────────────────────────────────
    if soh_path is None:
        soh_path = INPUTS_DIR / "soh.xlsx"
        print(f"[info] SOH path not specified — using default: {soh_path}")

    # ── Auto-detect order type ─────────────────────────────────────────────────
    if order_type is None:
        dow = today.dayofweek
        if dow == 2:       # Wednesday
            order_type = "WED_FRI"
        elif dow == 4:     # Friday
            order_type = "FRI_TUE"
        else:
            print(f"[error] Today is {today.strftime('%A')} — expected Wednesday or Friday.", file=sys.stderr)
            print("Use --order-type WED_FRI or FRI_TUE to override.", file=sys.stderr)
            sys.exit(1)

    print(f"[info] Order type: {order_type}")

    # ── Load base data ─────────────────────────────────────────────────────────
    print("[info] Loading sales snapshot…")
    all_sales = load_sales_from_snapshot()
    print(f"[info] Sales: {len(all_sales):,} rows, latest {all_sales['Date'].max().strftime('%d %b %Y')}")

    cal_df = load_calendar()
    model_data = load_model()
    if model_data:
        print(f"[info] Model loaded (trained {model_data.get('trained_on', 'unknown')})")
    else:
        print("[info] No model found — using EWMA fallback")

    # ── Cycle dates ────────────────────────────────────────────────────────────
    cycle_dates, delivery_date = get_cycle_dates(order_type, cal_df)
    if not cycle_dates:
        print("[error] No open trading days found for this cycle. Check the calendar.", file=sys.stderr)
        sys.exit(1)

    cycle_labels = [d.strftime("%a %d/%m") for d in cycle_dates]
    print(f"[info] Delivery: {delivery_date.strftime('%d %b %Y')}  Cycle: {' · '.join(cycle_labels)}")

    # ── Auto-detect specials paths from cycle dates ────────────────────────────
    if specials_path is None:
        first_cycle_date = cycle_dates[0].date()
        cycle_start = _specials_cycle_start(first_cycle_date)
        specials_path = find_specials_file(cycle_start)
        if specials_path:
            print(f"[info] Auto-detected specials: {specials_path.name}")
        else:
            print(f"[error] No specials file found for cycle starting {cycle_start} "
                  f"(searched: {INPUTS_DIR}/specials_{cycle_start}.[docx|doc])", file=sys.stderr)
            sys.exit(1)

    if specials_next_path is None and order_type == "FRI_TUE":
        _w2 = [d for d in cycle_dates if d.dayofweek >= 2]
        if _w2:
            next_cycle_start = _specials_cycle_start(_w2[0].date())
            specials_next_path = find_specials_file(next_cycle_start)
            if specials_next_path:
                print(f"[info] Auto-detected next-cycle specials: {specials_next_path.name}")
            else:
                print(f"[warn] No next-cycle specials file found for {next_cycle_start} — "
                      f"using current specials for all days")

    # ── Parse SOH ──────────────────────────────────────────────────────────────
    print(f"[info] Parsing SOH: {soh_path.name}")
    soh_map = parse_soh_file(soh_path)
    print(f"[info] SOH: {len(soh_map)} items with stock")

    # ── Parse Specials ─────────────────────────────────────────────────────────
    print(f"[info] Parsing specials: {specials_path.name}")
    cutoff = all_sales["Date"].max() - pd.Timedelta(weeks=ACTIVE_LOOKBACK_WEEKS)
    active_items = sorted(all_sales[all_sales["Date"] >= cutoff]["Name"].unique().tolist())

    specials_current = resolve_specials(specials_path, active_items)
    print(f"[info] Specials (current): {len(specials_current)} items on promotion")

    specials_next = specials_current  # fallback: use same list if no next file provided
    if specials_next_path is not None and specials_next_path.exists():
        specials_next = resolve_specials(specials_next_path, active_items)
        print(f"[info] Specials (next cycle): {len(specials_next)} items on promotion")
    elif order_type == "FRI_TUE":
        print("[warn] No next-cycle specials file provided — using current specials for all days")

    specials = specials_current  # used for order quantity calculations and Special label

    # ── Forecast ───────────────────────────────────────────────────────────────
    # FRI_TUE spans two specials weeks: Tue = current cycle, Wed+Thu = next cycle
    if order_type == "FRI_TUE":
        _w1_dates = [d for d in cycle_dates if d.dayofweek <= 1]   # Mon/Tue = current cycle
        _w2_dates = [d for d in cycle_dates if d.dayofweek >= 2]   # Wed/Thu = next cycle

        if _w1_dates and _w2_dates:
            fc_w1, lbl_w1, n1a, n1b = compute_forecast(all_sales, active_items, _w1_dates, specials_current, model_data)
            fc_w2, lbl_w2, n2a, n2b = compute_forecast(all_sales, active_items, _w2_dates, specials_next,    model_data)
            forecast_df = fc_w1.merge(fc_w2.drop(columns=["Total Forecast"]), on="Name", how="outer").fillna(0)
            c_labels    = lbl_w1 + lbl_w2
            forecast_df["Total Forecast"] = forecast_df[c_labels].sum(axis=1).round(1)
            n_lgbm = n1a + n2a
            n_ewma = n1b + n2b
        else:
            forecast_df, c_labels, n_lgbm, n_ewma = compute_forecast(
                all_sales, active_items, cycle_dates, specials_current, model_data
            )
    else:
        forecast_df, c_labels, n_lgbm, n_ewma = compute_forecast(
            all_sales, active_items, cycle_dates, specials_current, model_data
        )

    print(f"[info] Forecast: {n_lgbm} items via LightGBM, {n_ewma} via EWMA")

    # ── Assemble sheet_df ──────────────────────────────────────────────────────
    # SubDept from sales history (most recent for each item)
    subdept_map = (
        all_sales.sort_values("Date")
        .groupby("Name")["SubDept"]
        .last()
        .to_dict()
    ) if "SubDept" in all_sales.columns else {}

    revenue_map = (
        all_sales.groupby("Name")["Revenue"]
        .sum()
        .to_dict()
    ) if "Revenue" in all_sales.columns else {}

    soh_df   = pd.DataFrame(list(soh_map.items()), columns=["Name", "Sys_Stock"])
    sheet_df = forecast_df.merge(soh_df, on="Name", how="left")
    sheet_df["SubDept"]   = sheet_df["Name"].map(subdept_map).fillna("Other")
    sheet_df["Revenue"]   = sheet_df["Name"].map(revenue_map).fillna(0)
    sheet_df["is_system"] = sheet_df["Sys_Stock"].notna() & sheet_df["Name"].isin(soh_map)

    # Specials label
    specials_set = set(specials)
    sheet_df["Special"] = sheet_df["Name"].apply(lambda n: "S1" if n in specials_set else "")

    # ── Consolidation ──────────────────────────────────────────────────────────
    sheet_df = consolidate_forecast(sheet_df, c_labels, soh_map)

    # ── Filter to Freshlink only ───────────────────────────────────────────────
    sheet_df = sheet_df[
        sheet_df["Name"].apply(lambda n: item_belongs_to(n, "Freshlink"))
    ].copy().reset_index(drop=True)

    # ── Order quantities ───────────────────────────────────────────────────────
    sheet_df = calculate_order_quantities(
        sheet_df, all_sales, cycle_dates, delivery_date, cal_df, specials, model_data, c_labels
    )

    sheet_df = sheet_df.sort_values(["SubDept", "Name"]).reset_index(drop=True)
    print(f"[info] Sheet: {len(sheet_df)} items  "
          f"({int(sheet_df['is_system'].sum())} system-tracked, "
          f"{int((~sheet_df['is_system']).sum())} manual)")

    # ── Build Excel ────────────────────────────────────────────────────────────
    print("[info] Building Excel…")
    excel_bytes = build_excel(sheet_df, c_labels, cycle_dates, order_type)

    # ── Save output ────────────────────────────────────────────────────────────
    if output_path is None:
        OUTPUT_DIR.mkdir(exist_ok=True)
        fname = f"SCS_Freshlink_{date.today().strftime('%Y%m%d')}.xlsx"
        output_path = OUTPUT_DIR / fname

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(excel_bytes)
    print(f"[info] Saved: {output_path}")
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Headless order sheet generator")
    parser.add_argument("--soh",           default=None,  help="Path to SOH Excel/CSV file (default: 03_model/inputs/soh.xlsx)")
    parser.add_argument("--specials",      default=None,  help="Path to current-cycle specials .docx (default: auto-detect from cycle date)")
    parser.add_argument("--specials-next", default=None,  dest="specials_next",
                        help="Path to next-cycle specials .docx for FRI_TUE (default: auto-detect)")
    parser.add_argument("--output",        default=None,  help="Output Excel path (optional)")
    parser.add_argument("--order-type",    default=None,  dest="order_type",
                        choices=["WED_FRI", "FRI_TUE"],
                        help="Force order cycle (default: auto from weekday)")
    args = parser.parse_args()

    try:
        soh_path           = Path(args.soh)           if args.soh           else None
        specials_path      = Path(args.specials)      if args.specials      else None
        specials_next_path = Path(args.specials_next) if args.specials_next else None
        output_path        = Path(args.output)        if args.output        else None

        if soh_path and not soh_path.exists():
            print(f"[error] SOH file not found: {soh_path}", file=sys.stderr)
            sys.exit(1)
        if specials_path and not specials_path.exists():
            print(f"[error] Specials file not found: {specials_path}", file=sys.stderr)
            sys.exit(1)
        if specials_next_path and not specials_next_path.exists():
            print(f"[error] Next-cycle specials file not found: {specials_next_path}", file=sys.stderr)
            sys.exit(1)

        out = run(soh_path, specials_path, output_path, args.order_type, specials_next_path)
        # Print final path so calling scripts can capture it
        print(f"OUTPUT_PATH={out}")

    except Exception as e:
        import traceback
        print(f"[error] {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
