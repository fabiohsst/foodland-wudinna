"""
generate_price_updates.py — Foodland Wudinna Fruit & Veg
Pricing update tool: invoice cost → suggested sell prices at target GP%.

Usage:
    # Generate review sheet from an invoice CSV
    python generate_price_updates.py --invoice invoices/freshlink_20260407.csv

    # Apply approved prices from a completed review sheet
    python generate_price_updates.py --apply reviews/price_review_20260407.xlsx

Workflow:
    1. Run with --invoice to generate a review Excel.
    2. Review the Excel: confirm or override suggested prices, clear rows you want
       to skip, and set "Approve" to Y for each row you want to commit.
    3. Run with --apply to write approved prices back to item_price.csv.

Configuration (constants below):
    GP_TARGET          — target gross profit margin (0.40 = 40%)
    FLAG_THRESHOLD     — flag if cost OR sell price changes by more than this (0.15 = 15%)
    ITEM_PRICE_CSV     — path to the live price database
    INVOICE_MAPPING    — invoice description → POS item + unit conversion table
    PRICE_HISTORY_CSV  — time-series cost log; appended on every invoice run
"""

import argparse
import csv
import io
import re
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent

# Add project root to path so db.py is importable from the pricing/ subdirectory
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

GP_TARGET      = 0.40    # 40% gross profit margin target
FLAG_THRESHOLD = 0.15    # flag cost or sell price change ≥ 15%

INVOICE_MAPPING   = ROOT / "01_data" / "reference" / "invoice_item_mapping.csv"
SPECIALS_CSV      = ROOT / "01_data" / "operational" / "specials_this_week.csv"
REVIEWS_DIR       = ROOT / "pricing" / "reviews"

# CSV paths kept as fallbacks (used only if SQLite DB is unavailable)
ITEM_PRICE_CSV    = ROOT / "01_data" / "reference" / "item_price.csv"
PRICE_HISTORY_CSV = ROOT / "01_data" / "reference" / "price_history.csv"


# ── Helpers ───────────────────────────────────────────────────────────────────

def norm(s: str) -> str:
    """Collapse whitespace and strip — same normalisation as the rest of the project."""
    return re.sub(r"\s+", " ", str(s)).strip()


def round_to_x9(price: float) -> float:
    """
    Round up to the nearest price ending in 9 cents.
    Examples: 0.23 → 0.29 · 0.30 → 0.39 · 1.95 → 1.99 · 2.00 → 2.09 · 6.67 → 6.69
    """
    cents = round(price * 100)          # work in integer cents
    remainder = cents % 10
    if remainder == 9:
        return cents / 100
    return (cents + (9 - remainder)) / 100


def suggested_sell(cost_per_unit: float, gp_target: float = GP_TARGET) -> float:
    """
    Calculate the sell price required to achieve gp_target gross profit margin,
    then round up to the nearest X.X9 price.

    GP% = (Sell - Cost) / Sell  →  Sell = Cost / (1 - GP%)
    """
    if gp_target >= 1.0:
        raise ValueError("gp_target must be < 1.0 (e.g. 0.40 for 40%)")
    raw = cost_per_unit / (1.0 - gp_target)
    return round_to_x9(raw)


def pct_change(old: float, new: float) -> float | None:
    """Return % change from old to new, or None if old is zero/None."""
    if not old or old == 0:
        return None
    return (new - old) / old


# ── Data loaders ──────────────────────────────────────────────────────────────

def load_item_prices() -> dict[str, dict]:
    """
    Load item prices → {norm(name).upper(): {sell_price, cost_price, raw_name}}.
    Reads from SQLite (db.py); falls back to item_price.csv if DB is unavailable.
    """
    try:
        from db import load_item_price as _db_ip
        df = _db_ip()
        if not df.empty:
            prices = {}
            for _, row in df.iterrows():
                name = norm(str(row["Name"]))
                prices[name.upper()] = {
                    "sell_price": row.get("sell_price"),
                    "cost_price": row.get("cost_price"),
                    "raw_name":   name,
                }
            return prices
    except Exception:
        pass

    # CSV fallback
    prices = {}
    with open(ITEM_PRICE_CSV, newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            if len(row) < 4:
                continue
            name = norm(row[0])
            if name.lower() == "name":
                continue
            try:
                sell = float(row[2]) if row[2].strip() else None
                cost = float(row[3]) if row[3].strip() else None
            except ValueError:
                continue
            prices[name.upper()] = {"sell_price": sell, "cost_price": cost, "raw_name": name}
    return prices


def load_mapping() -> dict[str, dict]:
    """
    Load invoice description → POS item mapping.

    Reads from the DB (ref_invoice_mapping joined to dim_product) where possible;
    falls back to invoice_item_mapping.csv if the DB is unavailable.
    Returns {norm(invoice_description).upper(): mapping_dict}.
    """
    try:
        from db import load_invoice_mapping as _db_im
        df = _db_im()
        if not df.empty:
            mapping = {}
            for _, row in df.iterrows():
                if not row.get("invoice_description"):
                    continue
                key = norm(str(row["invoice_description"])).upper()
                mapping[key] = {
                    "pos_name":          norm(str(row["pos_name"])) if row.get("pos_name") else "",
                    "units_per_invoice": float(row["units_per_invoice"]) if row.get("units_per_invoice") else 1.0,
                    "sell_unit":         str(row.get("sell_unit", "each")).strip(),
                    "verified":          bool(row.get("verified", False)),
                    "notes":             str(row.get("notes", "")).strip(),
                }
            return mapping
    except Exception:
        pass

    # CSV fallback
    mapping = {}
    with open(INVOICE_MAPPING, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key = norm(row["invoice_description"]).upper()
            mapping[key] = {
                "pos_name":          norm(row["pos_name"]),
                "units_per_invoice": float(row["units_per_invoice"]),
                "sell_unit":         row["sell_unit"].strip(),
                "verified":          row.get("verified", "false").lower() == "true",
                "notes":             row.get("notes", "").strip(),
            }
    return mapping


def load_price_history() -> dict[str, dict]:
    """
    Load price history → {norm(pos_name).upper(): most_recent_row_dict}.
    Reads from SQLite (db.py); falls back to price_history.csv if unavailable.
    Prefers non-baseline entries; within those takes the most recent date.
    """
    try:
        from db import load_price_history as _db_ph
        df = _db_ph()
        if not df.empty:
            history: dict[str, list] = {}
            for _, row in df.iterrows():
                key = norm(str(row["pos_name"])).upper()
                history.setdefault(key, []).append(row)

            result = {}
            for key, rows in history.items():
                invoice_rows = [r for r in rows if r.get("source") != "item_price_baseline"]
                chosen = sorted(
                    invoice_rows or rows,
                    key=lambda r: str(r.get("date", "")),
                    reverse=True,
                )[0]
                try:
                    result[key] = {
                        "last_cost":   float(chosen["cost_per_unit"]),
                        "last_sell":   float(chosen["sell_price"]) if chosen.get("sell_price") else None,
                        "last_date":   str(chosen.get("date", "")),
                        "last_source": str(chosen.get("source", "")),
                    }
                except (ValueError, KeyError):
                    pass
            return result
    except Exception:
        pass

    # CSV fallback
    if not PRICE_HISTORY_CSV.exists():
        return {}
    history = {}
    with open(PRICE_HISTORY_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key = norm(row["pos_name"]).upper()
            history.setdefault(key, []).append(row)

    result = {}
    for key, rows in history.items():
        invoice_rows = [r for r in rows if r.get("source") != "item_price_baseline"]
        chosen = sorted(invoice_rows or rows, key=lambda r: r.get("date", ""), reverse=True)[0]
        try:
            result[key] = {
                "last_cost":   float(chosen["cost_per_unit"]),
                "last_sell":   float(chosen["sell_price"]) if chosen.get("sell_price") else None,
                "last_date":   chosen.get("date", ""),
                "last_source": chosen.get("source", ""),
            }
        except (ValueError, KeyError):
            pass
    return result


def load_current_specials() -> set[str]:
    """
    Load specials_this_week.csv → set of norm(pos_name) for items on special.
    Returns empty set if file not found.
    """
    specials: set[str] = set()
    if not SPECIALS_CSV.exists():
        return specials
    with open(SPECIALS_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            specials.add(norm(row.get("Name", "")).upper())
    return specials


# ── Invoice parsers ───────────────────────────────────────────────────────────

# Matches a Freshlink PDF invoice line, e.g.:
#   2 APPLES -PINK LADY -12 KG CTN $65.00 $130.00 FRE
#   1 APPLES -GRANNY SMITH -12 KG CTN $50.00 5% $47.50 FRE
_PDF_LINE = re.compile(
    r"^(\d+(?:\.\d+)?)\s+(.+?)\s+\$(\d+\.\d{2})\s*(?:([\d.]+)%)?\s+\$[\d,]+\.\d{2}\s+FRE"
)


def _norm_invoice_desc(s: str) -> str:
    """
    Normalise a Freshlink PDF description so it matches the mapping table.
    The PDF omits spaces around dashes ( -PINK LADY); the mapping uses ' - '.
    """
    s = re.sub(r"\s*-\s*", " - ", s.strip())
    return re.sub(r"\s+", " ", s).upper()


def parse_invoice_pdf(invoice_path: Path) -> list[dict]:
    """
    Parse a Freshlink PDF invoice using pdfplumber.

    Extracts invoice_no and date from the header, then matches each
    product line with the regex above.  Returns the same dict shape
    as parse_invoice() so process_invoice() needs no changes.
    """
    try:
        import pdfplumber
    except ImportError:
        print("ERROR: pdfplumber is not installed.  Run:  pip install pdfplumber")
        sys.exit(1)

    rows      = []
    invoice_no = ""
    date_str   = ""

    with pdfplumber.open(invoice_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()

                # Pull invoice number and date from the header (page 1)
                if not invoice_no:
                    m = re.search(r"Invoice No\s+(\d+)", line)
                    if m:
                        invoice_no = m.group(1)
                if not date_str:
                    m = re.search(r"Date\s+(\d{2}/\d{2}/\d{4})", line)
                    if m:
                        # Convert DD/MM/YYYY → YYYY-MM-DD
                        d, mo, y = m.group(1).split("/")
                        date_str = f"{y}-{mo}-{d}"

                # Match a product line
                m = _PDF_LINE.match(line)
                if not m:
                    continue

                qty_str, desc, price_str, disc_str = m.groups()
                unit_price = float(price_str)
                disc_pct   = float(disc_str) if disc_str else 0.0

                if unit_price <= 0:
                    continue  # lines with no price (e.g. some Peculiar Pick variants)

                net_price = unit_price * (1.0 - disc_pct / 100.0)

                rows.append({
                    "invoice_no":  invoice_no,
                    "date":        date_str,
                    "description": _norm_invoice_desc(desc),
                    "qty_ordered": float(qty_str),
                    "unit_price":  unit_price,
                    "disc_pct":    disc_pct,
                    "net_price":   net_price,
                })

    if not invoice_no:
        print("WARNING: Could not extract invoice number from PDF header.")
    if not date_str:
        date_str = date.today().isoformat()
        print(f"WARNING: Could not extract date from PDF — using today ({date_str}).")

    return rows


def parse_invoice(invoice_path: Path) -> list[dict]:
    """
    Parse an invoice file — auto-detects PDF vs CSV by file extension.

    PDF  → parse_invoice_pdf()  (Freshlink PDF format)
    CSV  → reads columns: invoice_no, date, description, qty_ordered,
                          unit_price, disc_pct
    Both return the same list-of-dicts shape for process_invoice().
    """
    if invoice_path.suffix.lower() == ".pdf":
        return parse_invoice_pdf(invoice_path)

    # ── CSV path ──────────────────────────────────────────────────────────────
    rows = []
    with open(invoice_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            try:
                unit_price = float(row["unit_price"])
                disc_pct   = float(row.get("disc_pct", 0) or 0)
                qty        = float(row.get("qty_ordered", 1) or 1)
            except (ValueError, KeyError):
                continue

            if unit_price <= 0:
                continue

            net_price = unit_price * (1.0 - disc_pct / 100.0)

            rows.append({
                "invoice_no":  row.get("invoice_no", "").strip(),
                "date":        row.get("date", "").strip(),
                "description": norm(row["description"]),
                "qty_ordered": qty,
                "unit_price":  unit_price,
                "disc_pct":    disc_pct,
                "net_price":   net_price,
            })
    return rows


# ── Core processor ────────────────────────────────────────────────────────────

def process_invoice(
    invoice_rows: list[dict],
    mapping:       dict[str, dict],
    item_prices:   dict[str, dict],
    history:       dict[str, dict],
    specials:      set[str],
    invoice_date:  str,
    invoice_no:    str,
) -> tuple[list[dict], list[str]]:
    """
    Match each invoice line to a POS item, calculate suggested sell price,
    and detect changes.

    Returns:
        results  — list of result dicts (one per matched invoice line)
        unmatched — list of invoice descriptions with no mapping entry
    """
    results   = []
    unmatched = []

    # Track which POS items have been processed to handle duplicates
    seen_pos: dict[str, dict] = {}

    for inv in invoice_rows:
        desc_key = inv["description"].upper()
        match    = mapping.get(desc_key)

        if match is None:
            unmatched.append(inv["description"])
            continue

        pos_name_norm = match["pos_name"].upper()
        units         = match["units_per_invoice"]
        cost_per_unit = inv["net_price"] / units

        # Skip items with zero or negative cost
        if cost_per_unit <= 0:
            continue

        # Lookup current price and history
        price_rec  = item_prices.get(pos_name_norm, {})
        hist_rec   = history.get(pos_name_norm, {})

        current_sell = price_rec.get("sell_price")
        current_cost = price_rec.get("cost_price")
        last_cost    = hist_rec.get("last_cost", current_cost)

        suggest = suggested_sell(cost_per_unit)

        cost_chg  = pct_change(last_cost, cost_per_unit)
        sell_chg  = pct_change(current_sell, suggest)

        on_special = pos_name_norm in specials

        # Build flag reason
        flags = []
        if cost_chg is not None and abs(cost_chg) >= FLAG_THRESHOLD:
            flags.append(f"Cost {'▲' if cost_chg > 0 else '▼'}{abs(cost_chg)*100:.0f}%")
        if sell_chg is not None and abs(sell_chg) >= FLAG_THRESHOLD:
            flags.append(f"Sell {'▲' if sell_chg > 0 else '▼'}{abs(sell_chg)*100:.0f}%")
        if not match["verified"]:
            flags.append("Unverified mapping")
        if on_special:
            flags.append("On special — skip")

        flag_str = " | ".join(flags) if flags else ""

        # Default approval: Y for clean unflagged items, N for flagged / on special
        approve = "N" if (on_special or flags) else "Y"

        # Actual GP% at suggested price (should be ≥ GP_TARGET)
        gp_actual = (suggest - cost_per_unit) / suggest if suggest > 0 else None

        row = {
            "Invoice Line":      inv["description"],
            "POS Item":          match["pos_name"],
            "Supplier":          "Freshlink",
            "Units/Invoice":     units,
            "Sell Unit":         match["sell_unit"],
            "Invoice Date":      invoice_date,
            "Unit Price (excl)": round(inv["unit_price"], 4),
            "Disc %":            inv["disc_pct"],
            "New Cost/Unit":     round(cost_per_unit, 4),
            "Prev Cost/Unit":    round(last_cost, 4) if last_cost else "",
            "Cost Δ%":           round(cost_chg * 100, 1) if cost_chg is not None else "",
            "Current Sell":      current_sell if current_sell else "",
            "Suggested Sell":    suggest,
            "Sell Δ%":           round(sell_chg * 100, 1) if sell_chg is not None else "",
            "GP% @ Suggested":   round(gp_actual * 100, 1) if gp_actual else "",
            "On Special":        "YES" if on_special else "",
            "Flag":              flag_str,
            "Approve":           approve,
            "Notes":             match["notes"],
        }

        # Handle duplicate POS items (e.g., two mushroom invoice lines → same POS item)
        # Keep the entry with the lower cost (more conservative pricing)
        if pos_name_norm in seen_pos:
            existing = seen_pos[pos_name_norm]
            if cost_per_unit < existing["New Cost/Unit"]:
                # Replace with cheaper option; note the duplicate
                row["Notes"] = f"Multiple invoice lines — using cheaper cost. {row['Notes']}"
                results[existing["_idx"]] = row
                seen_pos[pos_name_norm] = {**row, "_idx": existing["_idx"]}
            # else: keep existing, skip this one
        else:
            row["_idx"] = len(results)
            results.append(row)
            seen_pos[pos_name_norm] = {**row, "_idx": len(results) - 1}

    return results, unmatched


# ── Price history writer ───────────────────────────────────────────────────────

def append_price_history(results: list[dict], invoice_no: str, invoice_date: str) -> None:
    """Append this invoice's cost data to SQLite (idempotent per invoice_no + pos_name)."""
    rows = []
    for r in results:
        if not r.get("New Cost/Unit"):
            continue
        rows.append({
            "date":          invoice_date,
            "invoice_no":    invoice_no,
            "pos_name":      r["POS Item"],
            "cost_per_unit": r["New Cost/Unit"],
            "sell_price":    r.get("Current Sell") or None,
            "gp_pct":        r.get("GP% @ Suggested") or None,
            "source":        f"invoice_{invoice_no}",
        })

    if not rows:
        return

    try:
        from db import append_price_history as _db_aph
        inserted = _db_aph(rows)
        print(f"  Price history updated — {inserted} item(s) inserted (invoice {invoice_no})")
        return
    except Exception as e:
        print(f"  WARNING: SQLite write failed ({e}); falling back to CSV.")

    # CSV fallback
    existing = []
    if PRICE_HISTORY_CSV.exists():
        with open(PRICE_HISTORY_CSV, newline="", encoding="utf-8") as f:
            existing = list(csv.DictReader(f))
    existing = [r for r in existing if r.get("invoice_no") != invoice_no]
    all_rows = existing + [{k: (v if v is not None else "") for k, v in r.items()} for r in rows]
    with open(PRICE_HISTORY_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["date","invoice_no","pos_name","cost_per_unit","sell_price","gp_pct","source"])
        writer.writeheader()
        writer.writerows(all_rows)
    print(f"  Price history (CSV fallback) — {len(rows)} items written (invoice {invoice_no})")


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_review_excel(results: list[dict], unmatched: list[str],
                       invoice_path: Path, invoice_date: str) -> bytes:
    """Build the price review Excel workbook and return as bytes."""

    HEADER_BG    = "1A5276"
    FLAG_BG      = "FADBD8"    # light red — flagged items
    SPECIAL_BG   = "EBF5FB"   # light blue — on special (skip)
    APPROVE_BG   = "D5F5E3"   # light green — auto-approved
    NEUTRAL_BG   = "F8F9FA"   # alternate row
    WHITE_BG     = "FFFFFF"

    def fill(h):
        return PatternFill("solid", fgColor=h)

    thin = Border(
        left=Side("thin", color="CCCCCC"), right=Side("thin", color="CCCCCC"),
        top=Side("thin", color="CCCCCC"),  bottom=Side("thin", color="CCCCCC"),
    )
    center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Price Review"

    # ── Column definitions ────────────────────────────────────────────────────
    # (header_label, attribute_key, width, alignment, number_format)
    COLS = [
        ("Approve\n(Y/N)",         "Approve",           9,  center, "@"),
        ("Invoice Line",           "Invoice Line",      38, left_al, "@"),
        ("POS Item",               "POS Item",          32, left_al, "@"),
        ("Units/\nInvoice",        "Units/Invoice",     9,  center, "0.##"),
        ("Sell\nUnit",             "Sell Unit",         7,  center, "@"),
        ("Unit Price\n(ex disc)",  "Unit Price (excl)", 11, center, "0.00##"),
        ("Disc\n%",                "Disc %",            6,  center, "0.0"),
        ("New Cost\n/Unit",        "New Cost/Unit",     10, center, "0.0000"),
        ("Prev Cost\n/Unit",       "Prev Cost/Unit",    10, center, "0.0000"),
        ("Cost\nΔ%",               "Cost Δ%",           8,  center, '+0.0%;-0.0%;0.0%'),
        ("Current\nSell $",        "Current Sell",      10, center, "0.00"),
        ("Suggested\nSell $",      "Suggested Sell",    11, center, "0.00"),
        ("Sell\nΔ%",               "Sell Δ%",           8,  center, '+0.0%;-0.0%;0.0%'),
        ("GP%\n@ Suggested",       "GP% @ Suggested",   10, center, "0.0"),
        ("On\nSpecial",            "On Special",        8,  center, "@"),
        ("⚠ Flag",                 "Flag",              30, left_al, "@"),
        ("Notes",                  "Notes",             30, left_al, "@"),
    ]

    # ── Row 1: title ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    c = ws["A1"]
    c.value     = f"PRICE REVIEW — Foodland Wudinna | Invoice {invoice_path.stem} | {invoice_date}"
    c.font      = Font("Arial", 12, bold=True, color="FFFFFF")
    c.fill      = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: legend ─────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 14
    ws.merge_cells(f"A2:{get_column_letter(len(COLS))}2")
    ws["A2"].value = (
        "Green = auto-approved (no flags)   Red = flagged (review required)   "
        "Blue = on special (skipped)   Set Approve=Y to commit, N to skip."
    )
    ws["A2"].font      = Font("Arial", 8, italic=True, color="555555")
    ws["A2"].fill      = fill("EAF2FF")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: column headers ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 36
    hdr_font = Font("Arial", 9, bold=True, color="FFFFFF")
    for ci, (hdr, _, width, align, _) in enumerate(COLS, 1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = width
        c = ws.cell(3, ci, hdr)
        c.font      = hdr_font
        c.fill      = fill(HEADER_BG)
        c.alignment = center
        c.border    = thin

    # ── Body ──────────────────────────────────────────────────────────────────
    for row_idx, rec in enumerate(results, 4):
        ws.row_dimensions[row_idx].height = 15

        on_special = bool(rec.get("On Special"))
        flagged    = bool(rec.get("Flag") and not (on_special and rec["Flag"] == "On special — skip"))
        approved   = rec.get("Approve") == "Y"

        if on_special:
            row_bg = fill(SPECIAL_BG)
        elif flagged:
            row_bg = fill(FLAG_BG)
        elif approved:
            row_bg = fill(APPROVE_BG)
        else:
            row_bg = fill(NEUTRAL_BG if row_idx % 2 == 0 else WHITE_BG)

        for ci, (_, key, _, align, num_fmt) in enumerate(COLS, 1):
            val = rec.get(key, "")

            # Convert numeric-looking strings for proper formatting
            if key in ("Cost Δ%", "Sell Δ%") and val != "":
                try:
                    val = float(val) / 100   # format as percentage
                except (ValueError, TypeError):
                    pass

            c = ws.cell(row_idx, ci, val if val != "" else None)
            c.fill      = row_bg
            c.border    = thin
            c.alignment = align
            c.font      = Font("Arial", 9)
            if num_fmt != "@":
                c.number_format = num_fmt

        # Highlight Approve column
        approve_cell = ws.cell(row_idx, 1)
        if on_special:
            approve_cell.font = Font("Arial", 9, color="1A5276")
        elif approved:
            approve_cell.font = Font("Arial", 9, bold=True, color="1E8449")
        else:
            approve_cell.font = Font("Arial", 9, bold=True, color="C0392B")

    # ── Unmatched items section ────────────────────────────────────────────────
    if unmatched:
        um_row = len(results) + 5
        ws.merge_cells(f"A{um_row}:{get_column_letter(len(COLS))}{um_row}")
        c = ws.cell(um_row, 1, f"UNMATCHED INVOICE LINES ({len(unmatched)}) — add to invoice_item_mapping.csv")
        c.font = Font("Arial", 9, bold=True, color="FFFFFF")
        c.fill = fill("922B21")
        for ci in range(1, len(COLS) + 1):
            ws.cell(um_row, ci).border = thin

        for i, desc in enumerate(unmatched, um_row + 1):
            ws.merge_cells(f"A{i}:{get_column_letter(len(COLS))}{i}")
            c = ws.cell(i, 1, f"  {desc}")
            c.font   = Font("Arial", 9, italic=True, color="7B241C")
            c.fill   = fill("FADBD8")
            c.border = thin

    # ── Summary stats ─────────────────────────────────────────────────────────
    total      = len(results)
    auto_ok    = sum(1 for r in results if r.get("Approve") == "Y")
    flagged_n  = sum(1 for r in results if r.get("Flag") and "On special" not in r["Flag"])
    special_n  = sum(1 for r in results if r.get("On Special"))

    stat_row = len(results) + 5 + (len(unmatched) + 2 if unmatched else 0)
    ws.merge_cells(f"A{stat_row}:{get_column_letter(len(COLS))}{stat_row}")
    c = ws.cell(stat_row, 1,
        f"Summary: {total} items processed | {auto_ok} auto-approved | "
        f"{flagged_n} flagged for review | {special_n} on special (skipped) | "
        f"{len(unmatched)} unmatched"
    )
    c.font      = Font("Arial", 9, italic=True, color="444444")
    c.fill      = fill("F2F3F4")
    c.alignment = left_al

    ws.freeze_panes = "B4"
    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Apply approved prices ─────────────────────────────────────────────────────

def apply_approved_prices(review_path: Path) -> None:
    """
    Read a completed review Excel (Approve=Y rows) and update item_price.csv.
    Creates a timestamped backup of item_price.csv before writing.
    """
    print(f"\nApplying approved prices from: {review_path.name}")

    # Load the review sheet
    wb = load_workbook(review_path, data_only=True)
    ws = wb.active

    # Read header row (row 3)
    headers = [ws.cell(3, ci).value for ci in range(1, ws.max_column + 1)]
    try:
        col_approve  = headers.index("Approve\n(Y/N)") + 1
        col_pos      = headers.index("POS Item") + 1
        col_suggest  = headers.index("Suggested Sell $") + 1
        col_cost     = headers.index("New Cost\n/Unit") + 1
    except ValueError as e:
        print(f"ERROR: Could not find required column in review sheet: {e}")
        sys.exit(1)

    approved: dict[str, tuple[float, float]] = {}
    for ri in range(4, ws.max_row + 1):
        approve = ws.cell(ri, col_approve).value
        if str(approve).strip().upper() != "Y":
            continue
        pos_name = ws.cell(ri, col_pos).value
        suggest  = ws.cell(ri, col_suggest).value
        cost     = ws.cell(ri, col_cost).value
        if pos_name and suggest and cost:
            approved[norm(pos_name).upper()] = (float(suggest), float(cost))

    if not approved:
        print("No approved rows found (Approve=Y). Nothing updated.")
        return

    print(f"  {len(approved)} item(s) to update.")

    # Build upsert payload
    upsert_rows = []
    for name_key, (suggest, cost) in approved.items():
        upsert_rows.append({
            "Name":        name_key,
            "sell_price":  suggest,
            "cost_price":  round(cost, 4),
            "price_source": "invoice",
        })
        print(f"    ✓ {name_key} → sell ${suggest:.2f}  cost ${cost:.4f}")

    # Write to SQLite
    try:
        from db import upsert_item_prices as _db_uip
        written = _db_uip(upsert_rows)
        print(f"\n  Done — {written} item(s) updated in SQLite item_price table")
        return
    except Exception as e:
        print(f"  WARNING: SQLite write failed ({e}); falling back to CSV.")

    # CSV fallback — read, backup, update, write
    import shutil
    rows_raw: list[list[str]] = []
    with open(ITEM_PRICE_CSV, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows_raw = list(reader)

    backup_path = ITEM_PRICE_CSV.with_suffix(
        f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    )
    shutil.copy(ITEM_PRICE_CSV, backup_path)
    print(f"  Backup saved: {backup_path.name}")

    updated = 0
    for row in rows_raw:
        if not row:
            continue
        name_key = norm(row[0]).upper()
        if name_key in approved:
            suggest, cost = approved[name_key]
            row[1] = str(suggest)
            row[2] = str(suggest)
            row[3] = str(round(cost, 4))
            updated += 1

    with open(ITEM_PRICE_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows_raw)

    print(f"\n  Done — {updated} item(s) updated in item_price.csv (CSV fallback)")


# ── CLI entrypoint ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate price update review sheet from a delivery invoice."
    )
    parser.add_argument("--invoice", "-i", type=Path, default=None,
                        help="Path to invoice CSV (invoices/freshlink_YYYYMMDD.csv).")
    parser.add_argument("--apply", "-a", type=Path, default=None,
                        help="Apply approved prices from a completed review Excel.")
    parser.add_argument("--gp", type=float, default=GP_TARGET,
                        help=f"GP% target as decimal (default: {GP_TARGET}).")
    args = parser.parse_args()

    if args.apply:
        apply_approved_prices(args.apply)
        return

    if not args.invoice:
        parser.print_help()
        sys.exit(1)

    invoice_path = Path(args.invoice)
    if not invoice_path.exists():
        print(f"ERROR: Invoice file not found: {invoice_path}")
        sys.exit(1)

    print(f"\nPrice Update Generator — Foodland Wudinna Fruit & Veg")
    print(f"Invoice : {invoice_path.name}")
    print(f"GP target: {args.gp * 100:.0f}%")
    print(f"Flag threshold: ±{FLAG_THRESHOLD * 100:.0f}%\n")

    # Load all reference data
    print("Loading reference data...")
    item_prices = load_item_prices()
    mapping     = load_mapping()
    history     = load_price_history()
    specials    = load_current_specials()
    print(f"  {len(item_prices)} items in price DB  |  {len(mapping)} mapping entries  |  "
          f"{len(history)} items in history  |  {len(specials)} on special")

    # Parse invoice
    invoice_rows = parse_invoice(invoice_path)
    if not invoice_rows:
        print("ERROR: No valid rows found in invoice CSV.")
        sys.exit(1)

    # Use date and invoice_no from first row
    invoice_date = invoice_rows[0].get("date", date.today().isoformat())
    invoice_no   = invoice_rows[0].get("invoice_no", invoice_path.stem)

    print(f"  {len(invoice_rows)} invoice lines loaded (date: {invoice_date}, no: {invoice_no})")

    # Process
    results, unmatched = process_invoice(
        invoice_rows, mapping, item_prices, history, specials,
        invoice_date, invoice_no
    )

    auto_ok   = sum(1 for r in results if r.get("Approve") == "Y")
    flagged_n = sum(1 for r in results if r.get("Flag") and "On special" not in r.get("Flag",""))
    special_n = sum(1 for r in results if r.get("On Special"))

    print(f"\nResults:")
    print(f"  {len(results)} items matched")
    print(f"  {auto_ok} auto-approved (no flags)")
    print(f"  {flagged_n} flagged for review")
    print(f"  {special_n} on special (skipped)")
    print(f"  {len(unmatched)} unmatched invoice lines")

    if unmatched:
        print("\nUnmatched (add to invoice_item_mapping.csv):")
        for u in unmatched:
            print(f"  - {u}")

    # Append to price history
    append_price_history(results, invoice_no, invoice_date)

    # Build and save Excel
    REVIEWS_DIR.mkdir(parents=True, exist_ok=True)
    date_slug   = invoice_date.replace("/", "")
    output_path = REVIEWS_DIR / f"price_review_{date_slug}.xlsx"
    excel_bytes = build_review_excel(results, unmatched, invoice_path, invoice_date)
    output_path.write_bytes(excel_bytes)

    print(f"\n✅ Review sheet saved: {output_path}")
    print(f"   Review and set Approve=Y/N, then run:")
    print(f"   python generate_price_updates.py --apply {output_path}")


if __name__ == "__main__":
    main()
