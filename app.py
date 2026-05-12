"""
app.py — Fruit & Veg Order Sheet Generator
Foodland Wudinna

Run with:  streamlit run app.py
Launch via: double-click "Launch Order App.bat"
"""

import io
import re
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT        = Path(__file__).parent
SOH_DIR     = ROOT / "01_data/operational"
SOH_CSV     = ROOT / "01_data/operational/stock_on_hand_v2.csv"  # kept as fallback

import sys as _sys
if str(ROOT) not in _sys.path:
    _sys.path.insert(0, str(ROOT))
CALENDAR    = ROOT / "07_powerbi/data/dim_calendar.csv"
OUTPUT_DIR  = ROOT / "04_ordering"
MODEL_PKL    = ROOT / "03_model/demand_model.pkl"
FORECAST_LOG = ROOT / "03_model/forecast_log.csv"

LOOKBACK_WEEKS  = 8
EWMA_ALPHA      = 0.35   # fallback EWMA alpha (new items not in model)
SOH_MIN         = 0      # stock values below this are treated as zero / unreliable
SOH_MAX         = 300    # stock values above this are treated as data errors and excluded

# ── Supplier item groups ───────────────────────────────────────────────────────
# Items are assigned to a supplier by name prefix (case-insensitive).
# Freshlink = everything that does NOT match any named-supplier prefix.
SUPPLIER_PREFIXES: dict[str, list[str] | None] = {
    "Freshlink":     None,              # sentinel — assigned by exclusion
    "Bowlsome":      ["BOWLSOME", "COMM CO"],
    "Local Kitchen": ["L/C", "L/K", "LK "],
    "Simply Tasty":  ["S/TASTY"],
}
_ALL_NAMED_PREFIXES = [
    p.upper()
    for s, prefs in SUPPLIER_PREFIXES.items() if prefs
    for p in prefs
]

def item_belongs_to(item_name: str, supplier: str) -> bool:
    """Return True if item_name belongs to the given supplier."""
    name_up = norm(item_name).upper()
    prefs   = SUPPLIER_PREFIXES[supplier]
    if prefs is None:   # Freshlink = not claimed by any named supplier
        return not any(name_up.startswith(p) for p in _ALL_NAMED_PREFIXES)
    return any(name_up.startswith(p.upper()) for p in prefs)

# ── Product consolidation ──────────────────────────────────────────────────────
# Items sold under multiple barcodes (cuts/sizes) that must be ordered as a
# single unit. Each group defines:
#   display_name  — the label shown on the order sheet
#   unit          — "whole" (heads) or "kg" (weight)
#   variants      — {POS item name: conversion factor to base unit}
#
# Whole-item groups: factor = fraction of one whole head
#   e.g. CABBAGE HALF → 0.5 whole; CABBAGE QUARTER → 0.25 whole
#   Result is rounded up to the nearest whole for ordering.
#
# Weight groups: factor = kg per sold unit
#   e.g. WATERMELON SEEDLESS PER KG → 1.0 kg (already by weight)
#        SLICED WATERMELON          → 1.5 kg per piece (adjust if cut size changes)

CONSOLIDATION_GROUPS = {
    "Cabbage": {
        "display_name": "CABBAGE (order whole heads)",
        "unit": "whole",
        "variants": {
            "Cabbage Whole":   1.00,
            "Cabbage Half":    0.50,
            "Cabbage Quarter": 0.25,
        },
    },
    "Cabbage Red": {
        "display_name": "CABBAGE RED (order whole heads)",
        "unit": "whole",
        "variants": {
            "Cabbage Red Quarter": 0.25,
        },
    },
    "Cabbage Chinese": {
        "display_name": "CABBAGE CHINESE (order whole heads)",
        "unit": "whole",
        "variants": {
            "Cabbage Chinese Whole": 1.00,
            "Cabbage Chinese Half":  0.50,
        },
    },
    "Cauliflower": {
        "display_name": "CAULIFLOWER (order whole heads)",
        "unit": "whole",
        "variants": {
            "Cauliflower Per Each": 1.00,
            "Cauliflower Half":     0.50,
        },
    },
    "Celery": {
        "display_name": "CELERY (order whole stalks)",
        "unit": "whole",
        "variants": {
            "Celery Large": 1.00,
            "Celery Half":  0.50,
        },
    },
    "Watermelon": {
        "display_name": "WATERMELON (order kg)",
        "unit": "kg",
        "variants": {
            "Watermelon Seedless Per Kg": 1.00,   # sold by weight (kg)
            "Sliced Watermelon":          1.00,   # also sold by weight (kg)
        },
    },
    "Rockmelon": {
        "display_name": "ROCKMELON (order kg)",
        "unit": "kg",
        "variants": {
            "Rockmelon Per Kg":  1.00,   # sold by weight (kg)
            "Sliced Rockmelon":  1.00,   # also sold by weight (kg)
        },
    },
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def save_forecast_log(sheet_df: pd.DataFrame, cycle_labels: list,
                      order_type: str, delivery_date: pd.Timestamp) -> None:
    """
    Append this cycle's forecast to 03_model/forecast_log.csv.

    De-duplication key: (order_date, order_type).  If the same cycle was
    already logged (e.g. from a previous test run today), those rows are
    replaced so the log always reflects the most recent generation.

    Schema:
        order_date | order_type | delivery_date | item_name | subdept
        is_consolidated | predicted_qty | order_qty
    """
    order_date = date.today().isoformat()

    rows = []
    for _, r in sheet_df.iterrows():
        rows.append({
            "order_date":      order_date,
            "order_type":      order_type,
            "delivery_date":   delivery_date.date().isoformat(),
            "item_name":       r["Name"],
            "dept":            r.get("FRUIT & VEG"),
            "subdept":         r.get("SubDept", "Other"),
            "is_consolidated": bool(r.get("_consolidated", False)),
            "predicted_qty":   round(float(r["Total Forecast"]), 2),
            "order_qty":       (None if pd.isna(r.get("Order_Qty"))
                                else round(float(r["Order_Qty"]), 2)),
        })

    new_df = pd.DataFrame(rows)

    if FORECAST_LOG.exists():
        existing = pd.read_csv(FORECAST_LOG)
        # Drop any rows for the same order_date + order_type (keep latest run)
        keep_mask = ~(
            (existing["order_date"]  == order_date) &
            (existing["order_type"]  == order_type)
        )
        existing = existing[keep_mask]
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        FORECAST_LOG.parent.mkdir(parents=True, exist_ok=True)
        combined = new_df

    combined.to_csv(FORECAST_LOG, index=False)


def consolidate_forecast(forecast_df: pd.DataFrame, cycle_labels: list,
                         soh_map: dict) -> pd.DataFrame:
    """
    Merge multi-barcode variants into a single consolidated order row.

    For whole-item groups (cabbage, cauliflower, celery):
      - Each variant's daily forecast is multiplied by its fraction-of-whole factor
      - Summed to give total whole equivalents per day
      - Total Forecast is rounded UP to the nearest whole for ordering
      - SOH is also converted to whole equivalents and summed

    For weight groups (watermelon, rockmelon):
      - Each variant's daily forecast is multiplied by its kg-per-unit factor
      - Summed to give total kg required per day

    Variant rows are removed and replaced by a single consolidated row.
    Items in a group that are not present in the forecast are silently skipped.
    """
    df = forecast_df.copy()
    rows_to_add = []
    names_to_drop = set()

    for group_key, group in CONSOLIDATION_GROUPS.items():
        display   = group["display_name"]
        unit      = group["unit"]
        variants  = group["variants"]   # {display-case name: factor}

        # Match variants against the forecast — case-insensitive
        norm_variants = {norm(k).lower(): v for k, v in variants.items()}
        matched = df[df["Name"].apply(lambda x: norm(x).lower()).isin(norm_variants.keys())]

        if matched.empty:
            continue   # none of this group's items are in the forecast

        # Aggregate daily columns
        agg_day = {}
        for label in cycle_labels:
            total = 0.0
            for _, row in matched.iterrows():
                factor = norm_variants.get(norm(row["Name"]).lower(), 1.0)
                total += float(row.get(label, 0) or 0) * factor
            agg_day[label] = round(total, 2)

        raw_total = sum(agg_day.values())

        if unit == "whole":
            # Round up to nearest whole head for the order quantity
            order_total = float(np.ceil(raw_total))
        else:
            # kg — keep one decimal
            order_total = round(raw_total, 1)

        # Consolidated SOH in base units
        consolidated_soh = 0.0
        for variant_name, factor in norm_variants.items():
            # Find original POS name in soh_map (case-insensitive)
            for soh_name, soh_qty in soh_map.items():
                if norm(soh_name).lower() == variant_name:
                    consolidated_soh += soh_qty * factor
                    break
        consolidated_soh = round(consolidated_soh, 1) if consolidated_soh > 0 else np.nan

        # Use SubDept from the first matched variant
        subdept = matched.iloc[0].get("SubDept", "Other")

        consolidated_row = {
            "Name":           display,
            "Total Forecast": order_total,
            "SubDept":        subdept,
            "Revenue":        matched["Revenue"].sum() if "Revenue" in matched.columns else 0,
            "Sys_Stock":      consolidated_soh if not np.isnan(consolidated_soh) else np.nan,
            "is_system":      not np.isnan(consolidated_soh) if consolidated_soh else False,
            "_consolidated":  True,
            "_unit":          unit,
            "_raw_total":     raw_total,
        }
        for label in cycle_labels:
            consolidated_row[label] = agg_day[label]

        rows_to_add.append(consolidated_row)
        names_to_drop.update(matched["Name"].tolist())

    # Drop variant rows and append consolidated rows
    df = df[~df["Name"].isin(names_to_drop)].copy()
    df["_consolidated"] = False
    df["_unit"]         = "unit"
    df["_raw_total"]    = df["Total Forecast"]

    if rows_to_add:
        df = pd.concat([df, pd.DataFrame(rows_to_add)], ignore_index=True)

    return df

def get_cycle_dates(order_type: str, cal_df: pd.DataFrame, today=None):
    """
    Return (list of coverage dates, delivery date) for the chosen order cycle.

    Normal cycles:
      WED_FRI   — order Wednesday, delivery Friday AM,    covers Fri→Tue
      FRI_TUE   — order Friday,    delivery Tuesday AM,   covers Tue→Fri

    Holiday cycles (use when Friday delivery is a public holiday):
      HOL_TUE_WED — order Tuesday,  delivery Wednesday night, covers Thu→Tue
      HOL_THU_TUE — order Thursday, delivery Tuesday night,   covers Wed→Sat
    """
    if today is None:
        today = pd.Timestamp.today().normalize()

    cal = cal_df.copy()
    cal["date"] = pd.to_datetime(cal["date"])

    def _next_dow(base, target_dow):
        """Next occurrence of target_dow (0=Mon…6=Sun) strictly after base."""
        days = (target_dow - base.dayofweek) % 7 or 7
        return base + pd.Timedelta(days=days)

    def _open_between(start, end):
        mask = (cal["date"] >= start) & (cal["date"] <= end) & (cal["is_store_open"] == 1)
        return sorted(cal.loc[mask, "date"].tolist())

    if order_type == "WED_FRI":
        delivery  = _next_dow(today, 4)          # next Friday
        cycle_end = _next_dow(delivery, 0)       # following Monday → covers Fri · Sat · Mon
        return _open_between(delivery, cycle_end), delivery

    elif order_type == "FRI_TUE":
        delivery  = _next_dow(today, 1)          # next Tuesday
        cycle_end = _next_dow(delivery, 3)       # following Thursday → covers Tue · Wed · Thu
        return _open_between(delivery, cycle_end), delivery

    elif order_type == "HOL_TUE_WED":
        # Order Tuesday → delivery Wednesday night → stock available from Thursday
        order_day     = _next_dow(today, 1)      # next Tuesday
        delivery      = order_day + pd.Timedelta(days=1)   # Wednesday
        coverage_start= delivery  + pd.Timedelta(days=1)   # Thursday
        cycle_end     = _next_dow(coverage_start, 1)       # following Tuesday
        return _open_between(coverage_start, cycle_end), delivery

    elif order_type == "HOL_THU_TUE":
        # Order Thursday → delivery Tuesday night → stock available from Wednesday
        order_day     = _next_dow(today, 3)      # next Thursday
        delivery      = _next_dow(order_day, 1)             # following Tuesday
        coverage_start= delivery + pd.Timedelta(days=1)    # Wednesday
        cycle_end     = _next_dow(coverage_start, 5)        # following Saturday
        return _open_between(coverage_start, cycle_end), delivery

    else:
        raise ValueError(f"Unknown order_type: {order_type!r}")


@st.cache_data(show_spinner=False, ttl=300)
def load_sales():
    """Load Fruit & Veg sales history from SQLite, normalised."""
    from db import load_sales as _db_load
    df = _db_load()
    if df.empty:
        raise FileNotFoundError("No sales data found in foodland_data.db.")
    # Belt-and-suspenders: enforce F&V scope even if the DB query ever returns more
    if "Department" in df.columns:
        df = df[df["Department"] == "FRUIT & VEG"].copy()
    # Add legacy 'Quantity' alias used by ordering logic
    df["Quantity"] = df["Qty"]
    return df


@st.cache_data(show_spinner=False, ttl=300)
def load_calendar():
    cal = pd.read_csv(CALENDAR)
    cal["date"] = pd.to_datetime(cal["date_key"].astype(str), format="%Y%m%d")
    return cal


@st.cache_resource(show_spinner=False)
def load_model():
    """Load the LightGBM pkl. Returns None if not found."""
    if not MODEL_PKL.exists():
        return None
    try:
        from predict import load_model as _load
        return _load()
    except Exception:
        return None


def load_soh_csv() -> dict:
    """
    Load latest stock-on-hand snapshot → {name: stock} (system-tracked items only).

    Reads from SQLite first; falls back to stock_on_hand_v2.csv if DB is absent.

    Note: the DB stores the import filename in the 'source' field, not the string
    'system'. All DB entries are system-tracked data, so we filter only by the
    valid stock range (0–SOH_MAX) to exclude bogus negatives from POS carry-forward.
    """
    try:
        from db import load_stock_on_hand as _db_soh
        soh = _db_soh()
        if not soh.empty:
            soh = soh.rename(columns={"name": "Name", "stock": "Stock", "source": "Source"})
            soh["Name"]  = soh["Name"].apply(norm)
            soh["Stock"] = pd.to_numeric(soh["Stock"], errors="coerce").fillna(0)
            # DB source is the import filename, not 'system' — all DB entries are
            # system data; apply range filter only to exclude POS carry-forward negatives.
            mask = soh["Stock"].between(SOH_MIN, SOH_MAX)
            return dict(zip(soh.loc[mask, "Name"], soh.loc[mask, "Stock"].round(0).astype(int)))
    except Exception:
        pass

    # CSV fallback
    if not SOH_CSV.exists():
        return {}
    soh = pd.read_csv(SOH_CSV)
    soh.columns = soh.columns.str.strip()
    soh["Name"]  = soh["Name"].apply(norm)
    soh["Stock"] = pd.to_numeric(soh["Stock"], errors="coerce").fillna(0)
    # Honour a Source column if present (manual vs system); otherwise treat all as system.
    if "Source" in soh.columns:
        mask = (soh["Source"] == "system") & soh["Stock"].between(SOH_MIN, SOH_MAX)
    else:
        mask = soh["Stock"].between(SOH_MIN, SOH_MAX)
    return dict(zip(soh.loc[mask, "Name"], soh.loc[mask, "Stock"].round(0).astype(int)))


def parse_soh_upload(uploaded_file) -> tuple[dict | None, str | None]:
    """
    Parse an uploaded SOH file (CSV or Excel) → ({name: stock}, error_message).

    Accepts any export where there is a recognisable item-name column and a
    stock/quantity column.  Returns (None, error_msg) if parsing fails.
    """
    try:
        fname = uploaded_file.name.lower()
        if fname.endswith(".csv"):
            soh = pd.read_csv(uploaded_file)
        elif fname.endswith((".xlsx", ".xls")):
            # POS exports have several metadata rows before the real header.
            # Scan for the first row whose first cell is "GTIN" or whose cells
            # contain recognisable column names, then re-read from that row.
            import openpyxl, io
            raw_bytes = uploaded_file.read()
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
            ws = wb.active
            header_row = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                cells = [str(c).strip().lower() if c is not None else "" for c in row]
                if "gtin" in cells or "description" in cells:
                    header_row = i  # 0-indexed
                    break
            if header_row is None:
                header_row = 0   # fall back to default behaviour
            soh = pd.read_excel(io.BytesIO(raw_bytes), header=header_row)
        else:
            return None, "Unsupported file type — upload a CSV or Excel file."

        # Normalise column names: strip whitespace and collapse internal newlines/spaces
        soh.columns = [
            " ".join(str(c).split()) if c is not None else "" for c in soh.columns
        ]

        # Detect item-name column
        _name_candidates = ("name", "item", "item name", "description", "product", "product name")
        name_col = next(
            (c for c in soh.columns if c.lower() in _name_candidates),
            None,
        )
        # Detect stock column  (handles "Stock On Hand" with embedded newline, etc.)
        _stock_candidates = ("stock", "qty", "quantity", "on hand", "onhand", "soh",
                             "current stock", "stock on hand", "available")
        stock_col = next(
            (c for c in soh.columns if c.lower() in _stock_candidates),
            None,
        )

        if name_col is None:
            return None, (
                f"Could not find an item-name column. "
                f"Columns found: {', '.join(soh.columns.tolist())}. "
                "Rename the product column to 'Name' or 'Item' and re-upload."
            )
        if stock_col is None:
            return None, (
                f"Could not find a stock column. "
                f"Columns found: {', '.join(soh.columns.tolist())}. "
                "Rename the quantity column to 'Stock' or 'Qty' and re-upload."
            )

        soh = soh.rename(columns={name_col: "Name", stock_col: "Stock"})
        soh["Name"]  = soh["Name"].apply(norm)
        soh["Stock"] = pd.to_numeric(soh["Stock"], errors="coerce").fillna(0)

        # If a Source column exists honour it; otherwise treat everything as system data
        if "Source" in soh.columns:
            mask = (soh["Source"] == "system") & soh["Stock"].between(SOH_MIN, SOH_MAX)
        else:
            mask = soh["Stock"].between(SOH_MIN, SOH_MAX)

        result = dict(zip(
            soh.loc[mask, "Name"],
            soh.loc[mask, "Stock"].round(0).astype(int),
        ))
        return result, None

    except Exception as exc:
        return None, f"Error reading file: {exc}"


# ── Specials bulletin parsing helpers ─────────────────────────────────────────
SPECIALS_MAPPING_CSV = ROOT / "01_data/reference/specials_mapping.csv"

@st.cache_data(show_spinner=False)
def load_specials_mapping() -> pd.DataFrame:
    """Load bulletin_description → pos_name mapping table."""
    if not SPECIALS_MAPPING_CSV.exists():
        return pd.DataFrame(columns=["bulletin_description", "pos_name", "supplier", "verified", "notes"])
    df = pd.read_csv(SPECIALS_MAPPING_CSV)
    df.columns = df.columns.str.strip()
    return df


def _clean_bulletin_description(raw: str) -> str:
    """
    Strip noise from a bulletin description cell before matching.
    Removes: APN numbers, bold markers, size strings, 'New Season/Line', em-dashes.
    """
    s = raw
    s = re.sub(r"\*+[^*]+\*+", " ", s)                            # **New Line** etc.
    s = re.sub(r"\bAPN\s*[\d\s]+", " ", s, flags=re.IGNORECASE)   # APN 9315737007910
    s = re.sub(r"\b\d+(\.\d+)?\s*(g|kg|ml|l)\b", " ", s, flags=re.IGNORECASE)  # 400g, 1.5kg
    s = re.sub(r"\bPer\s+Carton\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b\d+\s*Per\s+Carton\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bNew\s+(Line|Season)\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bPre[-\s]?orders\s+required\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"[–—]", " ", s)    # em/en dash
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _extract_bulletin_descriptions(doc) -> list[str]:
    """
    Pull item descriptions from the specials bulletin table (column index 1 = Description).
    Skips header rows, blank rows, cost-only sub-rows, and supplier contact rows.
    Returns cleaned descriptions ready for matching.
    """
    _skip_patterns = re.compile(
        r"contact\s+phone|phone\s+number|\d{2}\s\d{4}\s\d{4}"
        r"|^j\s+l\s+king|metcash\s+item\s+no|^rainbow\s+fresh",
        re.IGNORECASE,
    )
    seen, results = set(), []
    for table in doc.tables:
        for row in table.rows:
            cells = row.cells
            if len(cells) < 2:
                continue
            raw = re.sub(r"\s+", " ", cells[1].text).strip()
            if not raw:
                continue
            lower = raw.lower()
            if lower in ("description", "size count", "desc", "item"):
                continue
            # Skip cost-only rows (e.g. "$2.45 per kg")
            if re.match(r"^\$?[\d,.]+\s*(per\s+kg|ea|pack|each)?$", lower):
                continue
            if _skip_patterns.search(lower):
                continue
            if len(raw) < 4:
                continue
            clean = _clean_bulletin_description(raw)
            if len(clean) < 4:
                continue
            if clean not in seen:
                seen.add(clean)
                results.append(clean)
    return results


def _find_libreoffice() -> str | None:
    """
    Locate the LibreOffice soffice executable across platforms.
    Returns the full path/command string, or None if not found.
    """
    import shutil, sys as _sys

    # Standard PATH lookup (works on Linux/Mac and Windows if soffice is in PATH)
    for cmd in ("soffice", "soffice.exe"):
        if shutil.which(cmd):
            return cmd

    # Windows — check the two common installation locations
    if _sys.platform == "win32":
        candidates = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for c in candidates:
            if Path(c).exists():
                return c

    return None


def parse_specials_bulletin(uploaded_file) -> tuple[list[str], str | None]:
    """
    Parse a Freshlink specials bulletin (.doc or .docx).
    Returns (descriptions, error_message).
    """
    import subprocess, tempfile
    from docx import Document as _DocxDoc

    fname = uploaded_file.name.lower()

    if fname.endswith(".docx"):
        try:
            doc = _DocxDoc(uploaded_file)
            return _extract_bulletin_descriptions(doc), None
        except Exception as e:
            return [], f"Could not read .docx: {e}"

    if fname.endswith(".doc"):
        soffice = _find_libreoffice()
        if soffice is None:
            return [], (
                "LibreOffice is required to open .doc files but was not found. "
                "Open the file in Word and save a copy as .docx, then upload that instead."
            )
        try:
            data = uploaded_file.read()
            with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as tmp:
                tmp.write(data)
                tmp_path = Path(tmp.name)
            out_dir  = tmp_path.parent
            out_docx = out_dir / (tmp_path.stem + ".docx")
            subprocess.run(
                [soffice, "--headless", "--convert-to", "docx",
                 "--outdir", str(out_dir), str(tmp_path)],
                capture_output=True, timeout=45,
            )
            tmp_path.unlink(missing_ok=True)
            if not out_docx.exists():
                return [], (
                    "LibreOffice could not convert the file. "
                    "Open it in Word and save as .docx, then upload that instead."
                )
            doc = _DocxDoc(out_docx)
            descs = _extract_bulletin_descriptions(doc)
            out_docx.unlink(missing_ok=True)
            return descs, None
        except Exception as e:
            return [], f"Error converting .doc: {e}"

    return [], f"Unsupported file type: {uploaded_file.name}"


def match_bulletin_to_pos(descriptions: list[str],
                           mapping_df: pd.DataFrame,
                           available_names: list[str]) -> tuple[list[str], list[str]]:
    """
    Match cleaned bulletin descriptions → POS names using specials_mapping.csv.

    Strategy (in order):
      1. Exact match on normalised text
      2. Token overlap ≥ 2 significant words from the mapping key appear in desc
      3. Fuzzy string similarity ≥ 0.62

    Returns (matched_pos_names, unmatched_descriptions).
    """
    import difflib

    if mapping_df.empty:
        return [], descriptions

    available_upper = {n.upper(): n for n in available_names}

    # Stop-words that should not count as significant matches
    _STOP = {
        "the", "and", "or", "of", "a", "an", "in", "on", "at", "to", "per",
        "bag", "pack", "punnet", "each", "kg", "g", "ml", "l", "size", "count",
        "punnets", "packs", "new", "season", "line", "fresh",
    }

    def _tokens(s: str) -> set[str]:
        return {w for w in re.sub(r"[^a-z0-9\s]", " ", s.lower()).split()
                if w not in _STOP and len(w) > 2}

    # Build normalised mapping table
    mdf = mapping_df.copy()
    mdf["_key"]    = mdf["bulletin_description"].str.upper().str.strip()
    mdf["_key"]    = mdf["_key"].apply(lambda s: re.sub(r"\s+", " ", s))
    mdf["_tokens"] = mdf["_key"].apply(lambda k: _tokens(k))

    matched, unmatched, seen_pos = [], [], set()

    def _add(pos_name: str) -> None:
        canon    = pos_name.strip()
        resolved = available_upper.get(canon.upper(), canon)
        if resolved not in seen_pos and resolved.upper() in available_upper:
            seen_pos.add(resolved)
            matched.append(resolved)

    for desc in descriptions:
        norm_key = re.sub(r"\s+", " ", desc.upper().strip())
        desc_toks = _tokens(desc)

        # 1. Exact match
        exact = mdf[mdf["_key"] == norm_key]
        if not exact.empty:
            _add(exact.iloc[0]["pos_name"])
            continue

        # 2. Token overlap — require ≥ 2 tokens from the mapping key in the description
        best_overlap, best_row = 0, None
        for _, mrow in mdf.iterrows():
            mk_toks = mrow["_tokens"]
            if len(mk_toks) == 0:
                continue
            overlap = len(mk_toks & desc_toks)
            # Also require the overlap to be at least half of the mapping key tokens
            if overlap >= 2 and overlap / len(mk_toks) >= 0.5:
                if overlap > best_overlap:
                    best_overlap = overlap
                    best_row = mrow
        if best_row is not None:
            _add(best_row["pos_name"])
            continue

        # 3. Fuzzy similarity
        keys  = mdf["_key"].tolist()
        close = difflib.get_close_matches(norm_key, keys, n=1, cutoff=0.62)
        if close:
            row = mdf[mdf["_key"] == close[0]].iloc[0]
            _add(row["pos_name"])
            continue

        unmatched.append(desc)

    return matched, unmatched


def ewma_forecast(sales: pd.DataFrame, items: list, cycle_dates: list, alpha: float = EWMA_ALPHA):
    """Simple EWMA fallback — used for items not covered by the LightGBM model."""
    sales = sales.copy()
    sales["dow"] = sales["Date"].dt.dayofweek
    cycle_labels = [d.strftime("%a %d/%m") for d in cycle_dates]

    forecast = pd.DataFrame({"Name": items})
    for tgt_date, label in zip(cycle_dates, cycle_labels):
        dow = tgt_date.dayofweek
        day_f = []
        for item in items:
            item_df  = sales[sales["Name"] == item]
            same_dow = item_df[item_df["dow"] == dow].sort_values("Date")
            qtys     = same_dow["Quantity"].values
            if len(qtys) == 0:
                day_f.append(round(float(item_df["Quantity"].mean()) if len(item_df) else 0.0, 1))
                continue
            lags    = qtys[-6:][::-1]
            weights = np.array([alpha ** i for i in range(len(lags))])
            weights /= weights.sum()
            day_f.append(round(float(np.dot(weights, lags)), 1))
        forecast[label] = day_f

    forecast["Total Forecast"] = forecast[cycle_labels].sum(axis=1).round(1)
    return forecast, cycle_labels


ACTIVE_LOOKBACK_WEEKS = 2   # items not sold in this window are excluded from forecast

def compute_forecast(all_sales: pd.DataFrame, active_items: list, cycle_dates: list,
                     specials: list, model_data) -> tuple:
    """
    Forecast demand for active_items over cycle_dates.

    Prefilter: only items sold in the last 2 weeks (or on special) are forecast.
    - Items the model knows → LightGBM prediction
    - New items (not in model) → EWMA fallback from recent sales

    Returns (forecast_df, cycle_labels, n_lgbm, n_ewma)
    """
    cycle_labels = [d.strftime("%a %d/%m") for d in cycle_dates]

    # Build the active set: sold in last 2 weeks OR flagged as special
    active_cutoff  = all_sales["Date"].max() - pd.Timedelta(weeks=ACTIVE_LOOKBACK_WEEKS)
    recently_sold  = set(all_sales[all_sales["Date"] >= active_cutoff]["Name"].unique())
    active_filter  = recently_sold | set(specials)

    # Apply filter to the requested item list
    filtered_items = [i for i in active_items if i in active_filter]

    # Split into model-covered and new items
    if model_data is not None:
        model_items = set(model_data["active_items"])
        lgbm_items  = [i for i in filtered_items if i in model_items]
        ewma_items  = [i for i in filtered_items if i not in model_items]
    else:
        lgbm_items = []
        ewma_items = filtered_items

    parts = []

    # LightGBM predictions — pass active_filter so predict_cycle skips stale items
    if lgbm_items:
        from predict import predict_cycle
        lgbm_df, _ = predict_cycle(
            cycle_dates, specials, model_data, all_sales,
            active_filter=active_filter,
        )
        lgbm_df = lgbm_df[lgbm_df["Name"].isin(lgbm_items)].copy()
        parts.append(lgbm_df)

    # EWMA fallback for new items
    if ewma_items:
        cutoff      = all_sales["Date"].max() - pd.Timedelta(weeks=LOOKBACK_WEEKS)
        recent_ewma = all_sales[all_sales["Date"] >= cutoff]
        ewma_df, _  = ewma_forecast(recent_ewma, ewma_items, cycle_dates)
        parts.append(ewma_df)

    if parts:
        forecast_df = pd.concat(parts, ignore_index=True)
    else:
        forecast_df = pd.DataFrame({"Name": filtered_items, "Total Forecast": 0.0})
        for label in cycle_labels:
            forecast_df[label] = 0.0

    return forecast_df, cycle_labels, len(lgbm_items), len(ewma_items)


def build_excel(sheet_df: pd.DataFrame, cycle_labels: list, cycle_dates: list,
                order_type: str) -> bytes:
    """
    Build the order workbook and return as bytes.

    Tab 1 — "Order Sheet": print-friendly summary — Item, Cycle Forecast,
            SOH on Hand, Order Qty (includes +1 day buffer). Designed to be
            handed to whoever fills in the physical order.

    Tab 2 — "Detail": full stock count sheet with day-by-day forecasts,
            pre-delivery forecast, projected SOH, and Order Qty (+1 day buffer).
    """
    n_days      = len(cycle_labels)

    # ── Shared style constants ─────────────────────────────────────────────────
    HEADER_BG    = "1A5276";  BLUE_ITEM    = "D6EAF8"
    GREY_ITEM    = "F2F3F4";  WHITE_ITEM   = "FFFFFF"
    FORECAST_BG  = "EBF5FB";  TOTAL_BG     = "D5E8D4"
    ORDER_BG     = "FCF3CF";  CONSOL_BG    = "E8DAEF"

    def fill(h):
        return PatternFill("solid", fgColor=h)

    thin = Border(
        left=Side("thin", color="CCCCCC"),   right=Side("thin", color="CCCCCC"),
        top=Side("thin", color="CCCCCC"),    bottom=Side("thin", color="CCCCCC"),
    )
    center  = Alignment(horizontal="center", vertical="center")
    left_al = Alignment(horizontal="left",   vertical="center")

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — Order Sheet (print-friendly)
    # ══════════════════════════════════════════════════════════════════════════
    wp = wb.active
    wp.title = "Order Sheet"

    wp.column_dimensions["A"].width = 4
    wp.column_dimensions["B"].width = 36
    wp.column_dimensions["C"].width = 7    # Special
    wp.column_dimensions["D"].width = 14   # Cycle Forecast
    wp.column_dimensions["E"].width = 13   # SOH on Hand
    wp.column_dimensions["F"].width = 13   # Order Qty

    cycle_str_p = " · ".join(d.strftime("%a %d %b") for d in cycle_dates)

    # Row 1 — Title
    wp.row_dimensions[1].height = 24
    wp.merge_cells("A1:F1")
    c = wp["A1"]
    c.value     = "FRUIT & VEG ORDER — Foodland Wudinna"
    c.font      = Font("Arial", 14, bold=True, color="FFFFFF")
    c.fill      = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2 — Cycle info
    wp.row_dimensions[2].height = 15
    wp.merge_cells("A2:F2")
    wp["A2"].value = (
        f"Delivery: {cycle_dates[0].strftime('%a %d %b %Y') if cycle_dates else '—'}  |  "
        f"Cycle covers: {cycle_str_p}  |  Order: {order_type}"
    )
    wp["A2"].font      = Font("Arial", 9, color="444444")
    wp["A2"].fill      = fill("EAF2FF")
    wp["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Row 3 — Buffer note
    wp.row_dimensions[3].height = 13
    wp.merge_cells("A3:F3")
    wp["A3"].value = (
        f"Order Qty includes +1 day safety buffer  "
        f"({n_days}-day cycle → buffer ≈ 1/{n_days} of cycle forecast per item)"
    )
    wp["A3"].font      = Font("Arial", 8, italic=True, color="555555")
    wp["A3"].fill      = fill("FDFEFE")
    wp["A3"].alignment = Alignment(horizontal="center", vertical="center")

    # Row 4 — Column headers
    wp.row_dimensions[4].height = 28
    p_hdr_font = Font("Arial", 10, bold=True, color="FFFFFF")
    for ci, hdr in enumerate(["#", "Item Name", "Special", "Cycle\nForecast", "SOH\non Hand", "Order Qty\n(+1 day)"], 1):
        c = wp.cell(4, ci, hdr)
        c.font      = p_hdr_font
        c.fill      = fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin

    # Body
    p_row = 5;  p_item = 0;  p_subdept = None

    for _, ir in sheet_df.iterrows():
        subdept   = ir["SubDept"]
        is_system = bool(ir["is_system"])
        is_consol = bool(ir.get("_consolidated", False))
        unit_lbl  = ir.get("_unit", "unit")

        # Sub-department separator
        if subdept != p_subdept:
            p_subdept = subdept
            wp.row_dimensions[p_row].height = 14
            wp.merge_cells(f"A{p_row}:F{p_row}")
            c = wp.cell(p_row, 1, subdept.upper())
            c.font      = Font("Arial", 9, bold=True, color="FFFFFF")
            c.fill      = fill(HEADER_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for ci in range(1, 7):
                wp.cell(p_row, ci).border = thin
            p_row += 1

        p_item += 1
        row_bg   = fill(CONSOL_BG) if is_consol else (fill(BLUE_ITEM) if is_system else
                   fill(GREY_ITEM) if p_row % 2 == 0 else fill(WHITE_ITEM))
        name_fnt = Font("Arial", 9, bold=(is_system or is_consol),
                        color="6C3483" if is_consol else "000000")

        wp.row_dimensions[p_row].height = 16

        # # col
        c = wp.cell(p_row, 1, p_item)
        c.font = Font("Arial", 8, color="888888");  c.fill = row_bg
        c.border = thin;  c.alignment = center

        # Item Name
        c = wp.cell(p_row, 2, ir["Name"])
        c.font = name_fnt;  c.fill = row_bg
        c.border = thin;  c.alignment = left_al

        # Special
        special_val = ir.get("Special", "")
        c = wp.cell(p_row, 3, special_val)
        if special_val:
            c.font = Font("Arial", 8, bold=True, color="7D6608")
            c.fill = fill("FEF9E7")
        else:
            c.font = Font("Arial", 8, color="CCCCCC")
            c.fill = row_bg
        c.border = thin;  c.alignment = center

        # Cycle Forecast
        fc_val = float(ir.get("_raw_total", ir["Total Forecast"]))
        fc_disp = int(np.ceil(fc_val)) if unit_lbl == "whole" else round(fc_val, 1)
        c = wp.cell(p_row, 4, fc_disp)
        c.font = Font("Arial", 9);  c.fill = fill(FORECAST_BG)
        c.border = thin;  c.alignment = center
        c.number_format = "0" if unit_lbl == "whole" else "0.0"

        # SOH on Hand
        if is_system or is_consol:
            soh_val = int(ir["Sys_Stock"]) if not pd.isna(ir.get("Sys_Stock")) else 0
            c = wp.cell(p_row, 5, soh_val)
            c.font = Font("Arial", 9, bold=True,
                          color="6C3483" if is_consol else "1A5276")
        else:
            c = wp.cell(p_row, 5, "")
            c.font = Font("Arial", 9)
        c.fill = row_bg;  c.border = thin;  c.alignment = center

        # Order Qty (buffered)
        buf_qty = ir.get("Order_Qty_Buf")
        has_buf = not pd.isna(buf_qty) if buf_qty is not None else False
        if has_buf:
            buf_disp = int(buf_qty) if unit_lbl == "whole" else round(float(buf_qty), 1)
            c = wp.cell(p_row, 6, buf_disp)
            c.font = Font("Arial", 10, bold=True,
                          color="C0392B" if buf_disp > 0 else "27AE60")
            c.number_format = "0" if unit_lbl == "whole" else "0.0"
        else:
            c = wp.cell(p_row, 6, "")
            c.font = Font("Arial", 9)
        c.fill = fill(ORDER_BG);  c.border = thin;  c.alignment = center

        p_row += 1

    wp.print_title_rows = "1:4"
    wp.page_setup.orientation = "portrait"
    wp.page_setup.paperSize   = 9   # A4
    wp.page_setup.fitToPage   = True
    wp.page_setup.fitToWidth  = 1
    wp.page_setup.fitToHeight = 0
    wp.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    wp.freeze_panes = "D5"

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — Detail (full stock count sheet)
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Detail")

    COL_NO      = 1;  COL_NAME    = 2;  COL_SPECIAL = 3
    COL_DAYS    = list(range(4, 4 + n_days))
    COL_SOH     = 4 + n_days
    COL_PRE_FC  = 5 + n_days
    COL_PROJ    = 6 + n_days
    COL_TOTAL   = 7 + n_days
    COL_ORDER   = 8 + n_days   # now shows buffered qty
    COL_NOTES   = 9 + n_days
    TOTAL_COLS  = COL_NOTES

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 7   # Special indicator
    for ci in COL_DAYS:
        ws.column_dimensions[get_column_letter(ci)].width = 8
    ws.column_dimensions[get_column_letter(COL_SOH)].width    = 8
    ws.column_dimensions[get_column_letter(COL_PRE_FC)].width = 8
    ws.column_dimensions[get_column_letter(COL_PROJ)].width   = 8
    ws.column_dimensions[get_column_letter(COL_TOTAL)].width  = 9
    ws.column_dimensions[get_column_letter(COL_ORDER)].width  = 9
    ws.column_dimensions[get_column_letter(COL_NOTES)].width  = 22

    # Row 1 — Title
    ws.row_dimensions[1].height = 24
    ws.merge_cells(f"A1:{get_column_letter(TOTAL_COLS)}1")
    c = ws["A1"]
    c.value = "STOCK COUNT SHEET — Foodland Wudinna | Fruit & Veg"
    c.font  = Font("Arial", 14, bold=True, color="FFFFFF")
    c.fill  = fill(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2 — Cycle info
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f"A2:{get_column_letter(TOTAL_COLS)}2")
    cycle_str = " | ".join(d.strftime("%a %d %b") for d in cycle_dates)
    ws["A2"].value = (
        f"Order date: {date.today().strftime('%d %b %Y')}  |  "
        f"Cycle: {cycle_str}  |  {order_type}"
    )
    ws["A2"].font      = Font("Arial", 9, color="444444")
    ws["A2"].fill      = fill("EAF2FF")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Row 3 — Legend
    ws.row_dimensions[3].height = 14
    ws.merge_cells(f"A3:{get_column_letter(TOTAL_COLS)}3")
    ws["A3"].value = (
        "Blue = system-tracked (SOH & Order Qty pre-filled)   "
        "White/Grey = count manually   "
        "Purple = consolidated multi-cut item (total in base unit)"
    )
    ws["A3"].font      = Font("Arial", 8, italic=True, color="555555")
    ws["A3"].fill      = fill("FDFEFE")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 4 — Column headers
    ws.row_dimensions[4].height = 32
    hdr_font = Font("Arial", 9, bold=True, color="FFFFFF")
    headers  = ["#", "Item Name", "Special"] + cycle_labels + ["SOH\nNow", "Fc to\nDeliv.", "SOH\nat Del.", "Total\nForecast", "Order Qty\n(+1 day)", "Notes"]
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(4, ci, hdr)
        c.font      = hdr_font
        c.fill      = fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin

    # Body
    row_num = 5;  item_num = 0;  current_subdept = None

    for _, ir in sheet_df.iterrows():
        subdept      = ir["SubDept"]
        is_system    = bool(ir["is_system"])
        is_consol    = bool(ir.get("_consolidated", False))
        unit_label   = ir.get("_unit", "unit")

        if subdept != current_subdept:
            current_subdept = subdept
            ws.row_dimensions[row_num].height = 16
            ws.merge_cells(f"A{row_num}:{get_column_letter(TOTAL_COLS)}{row_num}")
            c = ws.cell(row_num, 1, subdept.upper())
            c.font      = Font("Arial", 10, bold=True, color="FFFFFF")
            c.fill      = fill(HEADER_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for ci in range(1, TOTAL_COLS + 1):
                ws.cell(row_num, ci).border = thin
            row_num += 1

        item_num += 1

        if is_consol:
            row_bg      = fill(CONSOL_BG)
            fore_bg     = fill(CONSOL_BG)
            name_font   = Font("Arial", 9, bold=True, color="6C3483")
        elif is_system:
            row_bg      = fill(BLUE_ITEM)
            fore_bg     = fill(BLUE_ITEM)
            name_font   = Font("Arial", 9, bold=True)
        else:
            row_bg      = fill(GREY_ITEM) if row_num % 2 == 0 else fill(WHITE_ITEM)
            fore_bg     = fill(FORECAST_BG)
            name_font   = Font("Arial", 9)

        ws.row_dimensions[row_num].height = 16 if not is_consol else 18

        c = ws.cell(row_num, COL_NO, item_num)
        c.font = Font("Arial", 8, color="888888")
        c.fill = row_bg;  c.border = thin;  c.alignment = center

        c = ws.cell(row_num, COL_NAME, ir["Name"])
        c.font = name_font
        c.fill = row_bg;  c.border = thin;  c.alignment = left_al

        # Special indicator
        special_val = ir.get("Special", "")
        c = ws.cell(row_num, COL_SPECIAL, special_val)
        if special_val:
            c.font = Font("Arial", 8, bold=True, color="7D6608")
            c.fill = fill("FEF9E7")   # light amber — matches header scheme
        else:
            c.font = Font("Arial", 8, color="CCCCCC")
            c.fill = row_bg
        c.border = thin;  c.alignment = center

        for ci, label in zip(COL_DAYS, cycle_labels):
            val = round(float(ir.get(label, 0) or 0), 1)
            c = ws.cell(row_num, ci, val)
            c.font = Font("Arial", 9)
            c.fill = fore_bg
            c.border = thin;  c.alignment = center;  c.number_format = "0.0"

        # SOH Now
        if is_consol and not pd.isna(ir.get("Sys_Stock")):
            soh_v = round(float(ir["Sys_Stock"]), 1)
            c = ws.cell(row_num, COL_SOH, soh_v)
            c.font = Font("Arial", 9, bold=True, color="6C3483")
            c.fill = fill(CONSOL_BG)
        elif is_system:
            soh_v = int(ir["Sys_Stock"]) if not pd.isna(ir.get("Sys_Stock")) else 0
            c = ws.cell(row_num, COL_SOH, soh_v)
            c.font = Font("Arial", 9, bold=True, color="1A5276")
            c.fill = fill(BLUE_ITEM)
        else:
            c = ws.cell(row_num, COL_SOH, "")
            c.font = Font("Arial", 9);  c.fill = fill(WHITE_ITEM)
        c.border = thin;  c.alignment = center

        # Fc to Deliv — forecast sales from now until delivery arrives
        PRE_FC_BG = "EAF4FB"   # light teal
        pre_fc_v = float(ir.get("Pre_Del_Fc", 0) or 0)
        if pre_fc_v > 0 or (is_system or is_consol):
            c = ws.cell(row_num, COL_PRE_FC, round(pre_fc_v, 1))
            c.font = Font("Arial", 9, color="1A5276")
            c.fill = fill(PRE_FC_BG if not is_consol else CONSOL_BG)
        else:
            c = ws.cell(row_num, COL_PRE_FC, "")
            c.font = Font("Arial", 9);  c.fill = fill(WHITE_ITEM)
        c.border = thin;  c.alignment = center;  c.number_format = "0.0"

        # SOH at Delivery
        PROJ_BG = "FEF9E7"   # light amber
        has_proj = not pd.isna(ir.get("Proj_Stock"))
        if has_proj:
            proj_v = round(float(ir["Proj_Stock"]), 1)
            c = ws.cell(row_num, COL_PROJ, proj_v)
            proj_color = "6C3483" if is_consol else "784212"
            c.font = Font("Arial", 9, bold=True, color=proj_color)
            c.fill = fill(PROJ_BG if not is_consol else CONSOL_BG)
        else:
            c = ws.cell(row_num, COL_PROJ, "")
            c.font = Font("Arial", 9);  c.fill = fill(WHITE_ITEM)
        c.border = thin;  c.alignment = center

        # Total Forecast — net quantity needed: day_sum + Fc_to_Deliv − SOH_Now
        # Only shown for system/consolidated rows where SOH is known.
        net_v = ir.get("Net_Forecast")
        has_net = not pd.isna(net_v) if net_v is not None else False
        if has_net:
            net_display = float(net_v)
            c = ws.cell(row_num, COL_TOTAL, round(net_display, 1))
            net_color = "C0392B" if net_display > 0 else "27AE60"
            c.font = Font("Arial", 9, bold=True,
                          color="6C3483" if is_consol else net_color)
            c.fill = fill(TOTAL_BG)
            c.number_format = "0" if (is_consol and unit_label == "whole") else "0.0"
        else:
            c = ws.cell(row_num, COL_TOTAL, "")
            c.font = Font("Arial", 9);  c.fill = fill(WHITE_ITEM)
            c.number_format = "0.0"
        c.border = thin;  c.alignment = center

        # Order Qty (+1 day buffer)
        buf_v = ir.get("Order_Qty_Buf")
        has_order = not pd.isna(buf_v) if buf_v is not None else False
        if has_order:
            ord_display = int(buf_v) if unit_label == "whole" else round(float(buf_v), 1)
            c = ws.cell(row_num, COL_ORDER, ord_display)
            c.font = Font("Arial", 9, bold=True,
                          color="C0392B" if ord_display > 0 else "27AE60")
        else:
            c = ws.cell(row_num, COL_ORDER, "")
            c.font = Font("Arial", 9)
        c.fill = fill(ORDER_BG);  c.border = thin;  c.alignment = center

        # Notes — show unit hint for consolidated rows
        note_val = ""
        if is_consol:
            raw = ir.get("_raw_total", float(ir["Total Forecast"]))
            if unit_label == "whole":
                note_val = f"≈{raw:.1f} equiv → ↑ {int(np.ceil(raw))} whole"
            else:
                note_val = f"total {raw:.1f} kg"
        c = ws.cell(row_num, COL_NOTES, note_val)
        c.font = Font("Arial", 8, italic=is_consol, color="6C3483" if is_consol else "000000")
        c.fill = fill(CONSOL_BG if is_consol else WHITE_ITEM)
        c.border = thin;  c.alignment = left_al

        row_num += 1

    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.freeze_panes = "D5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Order Sheet Generator",
    page_icon="🥦",
    layout="centered",
)

st.title("🥦 Fruit & Veg — Order Sheet Generator")
st.caption("Foodland Wudinna")
st.divider()

# ── Load base data ─────────────────────────────────────────────────────────────
try:
    all_sales = load_sales()
    cal_df    = load_calendar()
except FileNotFoundError as e:
    st.error(f"Missing file: {e}")
    st.stop()

model_data = load_model()

# ── Data freshness check ───────────────────────────────────────────────────────
_data_max = all_sales["Date"].max()
_today    = pd.Timestamp.today().normalize()
_data_age = (_today - _data_max).days
if _data_age > 7:
    st.warning(
        f"⚠️ **Sales data is {_data_age} days old** (latest entry: {_data_max.strftime('%d %b %Y')}).  \n"
        "Active-item filtering and all forecasts are anchored to that date, not today.  \n"
        "Import a fresh POS export via the Import Panel, then click **↺ Refresh** below."
    )

_col_info, _col_btn = st.columns([4, 1])
with _col_info:
    st.caption(f"Sales data current to **{_data_max.strftime('%d %b %Y')}** · "
               f"Cache refreshes every 5 min or use the button →")
with _col_btn:
    if st.button("↺ Refresh", key="refresh_data"):
        st.cache_data.clear()
        st.cache_resource.clear()
        st.rerun()

# ── Section 1: Order Configuration ────────────────────────────────────────────
st.subheader("1 — Order Configuration")

# ── Supplier selector ─────────────────────────────────────────────────────────
supplier = st.selectbox(
    "Supplier",
    options=list(SUPPLIER_PREFIXES.keys()),
    index=0,
    help=(
        "Freshlink — full stock count sheet (standard process).  \n"
        "Bowlsome / Local Kitchen / Simply Tasty — email order format + audit sheet."
    ),
)
st.divider()

# ── Auto-detect holiday cycles ────────────────────────────────────────────────
# Check whether the next standard Friday or Tuesday delivery falls on a public holiday.
_today_norm  = pd.Timestamp.today().normalize()
_cal         = cal_df.copy()
_cal["date"] = pd.to_datetime(_cal["date"])

def _next_dow_date(base, dow):
    days = (dow - base.dayofweek) % 7 or 7
    return base + pd.Timedelta(days=days)

_next_fri = _next_dow_date(_today_norm, 4)
_next_tue = _next_dow_date(_today_norm, 1)
_fri_row  = _cal[_cal["date"] == _next_fri]
_tue_row  = _cal[_cal["date"] == _next_tue]
_fri_is_holiday = (not _fri_row.empty) and (_fri_row.iloc[0]["is_public_holiday"] == 1)
_tue_is_holiday = (not _tue_row.empty) and (_tue_row.iloc[0]["is_public_holiday"] == 1)

if _fri_is_holiday:
    _hol_name = _fri_row.iloc[0].get("holiday_name", "Public Holiday")
    st.warning(
        f"⚠️ **{_hol_name} ({_next_fri.strftime('%a %d %b')}) — Friday delivery is unavailable.**  \n"
        "Use **Holiday H1** (Tue order → Wed night delivery) instead of the normal Wednesday cycle.  \n"
        "After that stock runs out, switch to **Holiday H2** (Thu order → Tue night delivery)."
    )
if _tue_is_holiday:
    _hol_name = _tue_row.iloc[0].get("holiday_name", "Public Holiday")
    st.warning(
        f"⚠️ **{_hol_name} ({_next_tue.strftime('%a %d %b')}) — Tuesday delivery is unavailable.**  \n"
        "Contact your supplier for an alternative delivery day."
    )

# Smart default: auto-select the suggested cycle based on today
if _fri_is_holiday:
    _default_cycle = "HOL_TUE_WED"
elif pd.Timestamp.today().dayofweek == 4:   # Friday → normal FRI_TUE
    _default_cycle = "FRI_TUE"
else:
    _default_cycle = "WED_FRI"

_cycle_options = {
    "Wednesday → Friday  (normal)":           "WED_FRI",
    "Friday → Tuesday  (normal)":             "FRI_TUE",
    "Holiday H1 — Tue order, Wed night delivery": "HOL_TUE_WED",
    "Holiday H2 — Thu order, Tue night delivery": "HOL_THU_TUE",
}
_default_label = next(k for k, v in _cycle_options.items() if v == _default_cycle)

order_type_label = st.radio(
    "Order cycle",
    list(_cycle_options.keys()),
    index=list(_cycle_options.keys()).index(_default_label),
)
order_type = _cycle_options[order_type_label]

cycle_dates, delivery_date = get_cycle_dates(order_type, cal_df)

if not cycle_dates:
    st.warning("No store-open days found for this cycle. Check dim_calendar.csv or try another cycle.")
    st.stop()

cycle_labels = [d.strftime("%a %d/%m") for d in cycle_dates]

# FRI_TUE spans two specials bulletins (weeks run Wednesday → Tuesday):
#   S1 = current bulletin — Tue of cycle (final day of the current week)
#   S2 = next bulletin    — Wed and Thu  (first days of the new specials week)
_is_split_cycle = (order_type == "FRI_TUE")
if _is_split_cycle:
    _w1_dates = [d for d in cycle_dates if d.dayofweek <= 1]   # Mon/Tue
    _w2_dates = [d for d in cycle_dates if d.dayofweek >= 2]   # Wed–Sat
else:
    _w1_dates = cycle_dates
    _w2_dates = []

_cycle_note = {
    "WED_FRI":     "Wed order · Fri AM delivery · covers Fri, Sat, Mon",
    "FRI_TUE":     "Fri order · Tue AM delivery · covers Tue, Wed, Thu",
    "HOL_TUE_WED": "Tue order · Wed night delivery · stock from Thu",
    "HOL_THU_TUE": "Thu order · Tue night delivery · stock from Wed",
}[order_type]
st.info(
    f"📦 Delivery: **{delivery_date.strftime('%A %d %b')}**  ({_cycle_note})  \n"
    f"📅 Cycle covers: **{' · '.join(cycle_labels)}**  ({len(cycle_dates)} trading days)"
)

st.divider()

# ── Section 2: Specials ────────────────────────────────────────────────────────
st.subheader("2 — Specials This Cycle")

# Build candidate list for specials selector — all active items from recent sales
cutoff      = all_sales["Date"].max() - pd.Timedelta(weeks=LOOKBACK_WEEKS)
recent_all  = all_sales[all_sales["Date"] >= cutoff]

subdept_col = "Sub Department Name" if "Sub Department Name" in all_sales.columns else None
agg_spec    = {"Revenue": ("Revenue", "sum"), "Qty": ("Quantity", "sum")}
if subdept_col:
    agg_spec["SubDept"] = (
        subdept_col, lambda x: x.mode().iloc[0] if len(x) > 0 else "Other"
    )

active = (
    recent_all.groupby("Name").agg(**agg_spec)
    .sort_values("Revenue", ascending=False).reset_index()
)
if "SubDept" not in active.columns:
    active["SubDept"] = "Other"

# Filter items to the selected supplier so specials and forecast are scoped correctly
all_item_names = sorted(
    n for n in active["Name"].tolist()
    if item_belongs_to(n, supplier)
)

# ── Bulletin upload (optional) ────────────────────────────────────────────────
_specials_mapping = load_specials_mapping()

_bulletin_col, _manual_col = st.columns([1.6, 1], gap="large")

with _bulletin_col:
    st.markdown("**Upload specials bulletin** *(optional)*")
    _bulletin_file = st.file_uploader(
        "Freshlink specials sheet (.doc or .docx)",
        type=["doc", "docx"],
        label_visibility="collapsed",
        help=(
            "Upload the Freshlink weekly specials bulletin. "
            "Item descriptions are matched against the specials mapping table "
            "and pre-filled in the selector below. "
            "You can still add or remove items manually."
        ),
    )

with _manual_col:
    st.markdown("&nbsp;", unsafe_allow_html=True)
    st.caption(
        "Upload the weekly bulletin to pre-fill the specials list automatically, "
        "or use the selector below to add items manually. Both options work together."
    )

# ── Parse the bulletin if one was uploaded ────────────────────────────────────
_auto_specials_w1: list[str] = []
_auto_specials_w2: list[str] = []
_unmatched_descs:  list[str] = []

if _bulletin_file is not None:
    with st.spinner("Reading bulletin…"):
        _descs, _parse_err = parse_specials_bulletin(_bulletin_file)

    if _parse_err:
        st.error(f"Could not read bulletin: {_parse_err}")
    elif not _descs:
        st.warning("No item descriptions found in the uploaded file. Check the format.")
    else:
        _matched, _unmatched_descs = match_bulletin_to_pos(
            _descs, _specials_mapping, all_item_names
        )

        if _is_split_cycle:
            # Ask which bulletin week this sheet belongs to
            _week_choice = st.radio(
                "Which specials week is this bulletin for?",
                options=["S1 — Current bulletin", "S2 — Next bulletin", "Both weeks"],
                horizontal=True,
                help="S1 = current bulletin (Tuesday). S2 = next bulletin (Wed + Thu).",
            )
            if "S1" in _week_choice and "Both" not in _week_choice:
                _auto_specials_w1 = _matched
            elif "S2" in _week_choice and "Both" not in _week_choice:
                _auto_specials_w2 = _matched
            else:
                _auto_specials_w1 = _matched
                _auto_specials_w2 = _matched
        else:
            _auto_specials_w1 = _matched

        # Show a compact summary of matched vs total items extracted
        if _matched:
            st.success(
                f"✅ {len(_matched)} of {len(_descs)} item(s) from the bulletin "
                f"were matched to POS names and pre-filled below."
            )

st.markdown("---")

# ── Specials multiselect(s) — pre-filled from bulletin, editable manually ─────
if _is_split_cycle:
    # FRI_TUE spans two specials weeks — show a separate multiselect per week
    _w1_label_str = " · ".join(d.strftime("%a %d/%m") for d in _w1_dates)
    _w2_label_str = " · ".join(d.strftime("%a %d/%m") for d in _w2_dates)
    st.info(
        f"This cycle spans two specials bulletins.  \n"
        f"**S1 — current bulletin:** {_w1_label_str}  \n"
        f"**S2 — next bulletin:** {_w2_label_str}"
    )
    specials_w1 = st.multiselect(
        f"S1 — Current bulletin specials ({_w1_label_str})",
        options=all_item_names,
        default=[x for x in _auto_specials_w1 if x in all_item_names],
        placeholder="Start typing an item name…",
        key="specials_w1",
        help="Items on special during the current bulletin (Tuesday only for this cycle).",
    )
    specials_w2 = st.multiselect(
        f"S2 — Next bulletin specials ({_w2_label_str})",
        options=all_item_names,
        default=[x for x in _auto_specials_w2 if x in all_item_names],
        placeholder="Start typing an item name…",
        key="specials_w2",
        help="Items on special during next week's bulletin (Wednesday and Thursday).",
    )
    specials = list(set(specials_w1) | set(specials_w2))
    if specials:
        st.caption(
            f"🏷️ S1: {len(specials_w1)} item(s) · S2: {len(specials_w2)} item(s) "
            f"· {len(specials)} unique across cycle"
        )
    else:
        st.caption("No specials selected — all items forecast at normal demand.")
else:
    specials_w1 = st.multiselect(
        "Items on special this cycle",
        options=all_item_names,
        default=[x for x in _auto_specials_w1 if x in all_item_names],
        placeholder="Start typing an item name…",
        help="Items flagged here get cycle_on_special=1 in the LightGBM model, boosting their forecast.",
    )
    specials_w2 = []
    specials    = specials_w1
    if specials:
        st.caption(f"🏷️ {len(specials)} item(s) on special: {', '.join(specials)}")
    else:
        st.caption("No specials selected — all items forecast at normal demand.")

# ── Unmatched bulletin items — always visible after upload ────────────────────
if _unmatched_descs:
    st.warning(
        f"⚠️ **{len(_unmatched_descs)} item(s) from the bulletin could not be matched "
        f"to a POS name** — add them manually using the selector above if they apply to this cycle."
    )
    _um_lines = "\n".join(f"- {d}" for d in _unmatched_descs)
    st.markdown(_um_lines)
    st.caption(
        "To auto-match these next time, add them to "
        "`01_data/reference/specials_mapping.csv` with the correct POS name."
    )

st.divider()

# ── Section 3: Stock on Hand ───────────────────────────────────────────────────
st.subheader("3 — Stock on Hand (SOH)")

soh_file = st.file_uploader(
    "Upload your SOH export",
    type=["csv", "xlsx", "xls"],
    help=(
        "Download a stock-on-hand report from your POS system and upload it here. "
        "The file needs a product-name column (Name / Item / Description) and a "
        "stock column (Stock / Qty / On Hand). "
        "Items with stock ≥ 1 will have SOH and Order Qty pre-filled (blue rows)."
    ),
)

_soh_source = None   # tracks how soh_map was populated (for the UI caption)

if soh_file is not None:
    _parsed, _err = parse_soh_upload(soh_file)
    if _err:
        st.error(f"⚠️ Could not parse SOH file — {_err}")
        soh_map   = {}
        _soh_source = None
    else:
        soh_map     = _parsed
        _soh_source = "upload"
        st.success(
            f"✅ SOH loaded from **{soh_file.name}** — "
            f"{len(soh_map)} item(s) with stock ≥ 1 will have Order Qty pre-filled."
        )
else:
    # Fall back to folder file if present
    soh_map = load_soh_csv()
    if soh_map:
        _soh_source = "folder"
        st.caption(
            f"Using SOH from folder file (`stock_on_hand_v2.csv`) — "
            f"{len(soh_map)} item(s).  "
            "Upload a fresh export above to override."
        )
    else:
        _soh_source = None
        st.caption(
            "No SOH file uploaded and no `stock_on_hand_v2.csv` found.  "
            "SOH and Order Qty columns will be left blank for manual entry."
        )

st.divider()

# ── Build forecast ─────────────────────────────────────────────────────────────
with st.spinner("Computing forecast…"):
    if _is_split_cycle and _w1_dates and _w2_dates:
        # FRI_TUE: run separately for each specials week so cycle_on_special
        # applies to the correct days per bulletin.
        fc_w1, lbl_w1, n1a, n1b = compute_forecast(
            all_sales, active["Name"].tolist(), _w1_dates, specials_w1, model_data
        )
        fc_w2, lbl_w2, n2a, n2b = compute_forecast(
            all_sales, active["Name"].tolist(), _w2_dates, specials_w2, model_data
        )
        # Outer join on Name — missing days in either week default to 0
        forecast_df = fc_w1.merge(
            fc_w2.drop(columns=["Total Forecast"]), on="Name", how="outer"
        ).fillna(0)
        c_labels = lbl_w1 + lbl_w2
        forecast_df["Total Forecast"] = forecast_df[c_labels].sum(axis=1).round(1)
        n_lgbm = n1a + n2a
        n_ewma = n1b + n2b
    else:
        forecast_df, c_labels, n_lgbm, n_ewma = compute_forecast(
            all_sales,
            active["Name"].tolist(),
            cycle_dates,
            specials,
            model_data,
        )

# Assemble sheet_df — merge SOH and SubDept before consolidation
soh_df   = pd.DataFrame(list(soh_map.items()), columns=["Name", "Sys_Stock"])
sheet_df = (
    forecast_df
    .merge(soh_df, on="Name", how="left")
    .merge(active[["Name", "SubDept", "Revenue"]], on="Name", how="left")
)
sheet_df["SubDept"]   = sheet_df["SubDept"].fillna("Other")
sheet_df["is_system"] = sheet_df["Sys_Stock"].notna() & sheet_df["Name"].isin(soh_map)

# ── Apply product consolidation ────────────────────────────────────────────────
sheet_df = consolidate_forecast(sheet_df, c_labels, soh_map)

# ── Filter to selected supplier ────────────────────────────────────────────────
sheet_df = sheet_df[
    sheet_df["Name"].apply(lambda n: item_belongs_to(n, supplier))
].copy().reset_index(drop=True)

# ── Special cycle indicator ────────────────────────────────────────────────────
# Marks each item with which specials week it appears in: S1, S2, or S1+S2.
# For single-week cycles (WED_FRI, HOL variants), all specials are labelled S1.
_specials_w1_set = set(specials_w1)
_specials_w2_set = set(specials_w2)

def _special_label(name: str) -> str:
    in_w1 = name in _specials_w1_set
    in_w2 = name in _specials_w2_set
    if in_w1 and in_w2:
        return "S1+S2"
    elif in_w1:
        return "S1"
    elif in_w2:
        return "S2"
    return ""

sheet_df["Special"] = sheet_df["Name"].apply(_special_label)

# ── Pre-delivery stock depletion ───────────────────────────────────────────────
# Forecast demand for each open trading day between placing the order and delivery.
# Apply per-day weights so the projected SOH at delivery is accurate:
#   - Order day (today, after 12:00): weight 0.4 — roughly 40% of a full day remains
#   - All other open days (including Saturday): weight 1.0
#
# Saturday does NOT get a separate hour-fraction weight. The model already
# predicts Saturday quantities from real Saturday sales history, so its
# shorter trading hours (08:30–12:00) are baked into the forecast output.
# Applying an extra 0.368 multiplier would double-count the effect.
# The 0.4 order-day weight is the only external adjustment needed — the model
# has no knowledge of what time of day the order is placed.

_ORDER_DAY_WEIGHT = 0.4

# Get open trading days from today up to (not including) delivery
_cal_pre = cal_df.copy()
_cal_pre["date"] = pd.to_datetime(_cal_pre["date"])
_pre_mask = (
    (_cal_pre["date"] >= _today_norm) &
    (_cal_pre["date"] <  delivery_date) &
    (_cal_pre["is_store_open"] == 1)
)
_pre_delivery_dates = sorted(_cal_pre.loc[_pre_mask, "date"].tolist())

# Run the model for the pre-delivery window using the same items and specials
if _pre_delivery_dates:
    _pre_fc, _pre_labels, _, _ = compute_forecast(
        all_sales, active["Name"].tolist(), _pre_delivery_dates, specials, model_data
    )

    # Build weight per day label
    _day_weights = {}
    for d, lbl in zip(_pre_delivery_dates, _pre_labels):
        if d == _today_norm:
            _day_weights[lbl] = _ORDER_DAY_WEIGHT
        else:
            _day_weights[lbl] = 1.0

    # Weighted sum → per-item depletion estimate
    _dep_total = pd.Series(0.0, index=_pre_fc.index)
    for lbl, w in _day_weights.items():
        if lbl in _pre_fc.columns:
            _dep_total += _pre_fc[lbl].fillna(0) * w
    _pre_fc["_depletion"] = _dep_total
    _depletion_map = dict(zip(_pre_fc["Name"], _pre_fc["_depletion"]))
else:
    _depletion_map = {}   # holiday cycles or no gap — no depletion

def _proj_stock(row):
    """Projected SOH at delivery = Sys_Stock minus forecast-based pre-delivery depletion."""
    if pd.isna(row.get("Sys_Stock")):
        return np.nan
    depletion = _depletion_map.get(row["Name"], 0.0)
    return max(0.0, float(row["Sys_Stock"]) - depletion)

sheet_df["Proj_Stock"]  = sheet_df.apply(_proj_stock, axis=1)
sheet_df["Pre_Del_Fc"]  = sheet_df["Name"].map(_depletion_map).fillna(0.0).round(1)

def calc_order(row):
    if row.get("_consolidated") and not pd.isna(row.get("Proj_Stock")):
        # Consolidated row: use projected stock at delivery
        proj = float(row["Proj_Stock"])
        if row.get("_unit") == "whole":
            return max(0, int(np.ceil(row["_raw_total"])) - int(np.floor(proj)))
        else:
            return max(0, round(row["_raw_total"] - proj, 1))
    elif row.get("_consolidated"):
        return np.nan   # no SOH — buyer fills manually
    elif not row["is_system"]:
        return np.nan
    else:
        proj = float(row["Proj_Stock"]) if not pd.isna(row.get("Proj_Stock")) else 0
        return max(0, round(row["Total Forecast"] - proj))

sheet_df["Order_Qty"] = sheet_df.apply(calc_order, axis=1)

def calc_net_forecast(row):
    """
    Net quantity needed = day_sum + Fc_to_Deliv − SOH_Now.
    This is the raw order requirement before the max(0) floor.
    Blank for manual rows (no SOH known).
    """
    if pd.isna(row.get("Sys_Stock")):
        return np.nan
    soh     = float(row["Sys_Stock"])
    pre_del = float(row.get("Pre_Del_Fc", 0) or 0)
    raw     = float(row.get("_raw_total", row["Total Forecast"]))  # use raw for consolidated
    return round(raw + pre_del - soh, 1)

sheet_df["Net_Forecast"] = sheet_df.apply(calc_net_forecast, axis=1)

# ── Extra day buffer — safety stock to keep on shelf at end of cycle ──────────
# One extra day's average demand is added on top of the base Order_Qty.
# For "whole" items (heads, each) the buffer is rounded up to the nearest whole.
_n_cycle_days = len(cycle_dates)

def calc_buffer(row):
    raw = float(row.get("_raw_total", row["Total Forecast"]))
    buf = raw / max(_n_cycle_days, 1)
    if row.get("_unit") == "whole":
        return max(1, int(np.ceil(buf)))
    return round(buf, 1)

sheet_df["_buffer"] = sheet_df.apply(calc_buffer, axis=1)

sheet_df["Order_Qty_Buf"] = sheet_df.apply(
    lambda r: np.nan if pd.isna(r["Order_Qty"])
    else (int(r["Order_Qty"]) + int(r["_buffer"]) if r.get("_unit") == "whole"
          else round(float(r["Order_Qty"]) + float(r["_buffer"]), 1)),
    axis=1
)

sheet_df = sheet_df.sort_values(["SubDept", "Name"], ascending=[True, True]).reset_index(drop=True)

# ── Log this forecast (overwrites same order_date + order_type if re-run) ──────
save_forecast_log(sheet_df, c_labels, order_type, delivery_date)

# ── Section 4: Forecast Preview ───────────────────────────────────────────────
st.subheader("4 — Forecast Preview")

sys_count = int(sheet_df["is_system"].sum())
man_count = int((~sheet_df["is_system"]).sum())
tot_units = float(sheet_df["Total Forecast"].sum())
pre_calc  = float(sheet_df.loc[sheet_df["is_system"], "Order_Qty"].sum())

col1, col2, col3, col4 = st.columns(4)
col1.metric("Total Items",    len(sheet_df))
col2.metric("System-tracked", sys_count, help="SOH & Order Qty pre-filled (blue rows)")
col3.metric("Manual count",   man_count, help="SOH & Order Qty left blank (white/grey rows)")
col4.metric("Total Forecast", f"{tot_units:,.0f} units")

# Model indicator
if model_data is not None:
    trained_on = model_data.get("trained_on", "unknown")
    if n_ewma > 0:
        st.caption(
            f"Model: **LightGBM** ({n_lgbm} items) + EWMA fallback ({n_ewma} new items) · "
            f"Trained: {trained_on}"
        )
    else:
        st.caption(f"Model: **LightGBM** ({n_lgbm} items) · Trained: {trained_on}")
else:
    st.caption("Model: **EWMA** (LightGBM model not found — run FruitVeg_Demand_Forecast.ipynb to train)")

st.caption(f"Pre-calculated order qty for system items: **{pre_calc:.0f} units** — manual items require a physical count.")

# Sub-dept summary
dept_summary = (
    sheet_df.groupby("SubDept")
    .agg(Items=("Name", "count"), Forecast=("Total Forecast", "sum"))
    .sort_values("Forecast", ascending=False)
    .reset_index()
)
dept_summary["Forecast"] = dept_summary["Forecast"].round(1)
dept_summary.columns = ["Sub-Department", "Items", "Total Forecast (units)"]
st.dataframe(dept_summary, use_container_width=True, hide_index=True)

with st.expander("Show all items"):
    preview_cols = ["Name", "Special", "SubDept"] + c_labels
    if sys_count > 0:
        preview_cols += ["Sys_Stock", "Pre_Del_Fc", "Proj_Stock", "Net_Forecast", "Order_Qty"]
    else:
        preview_cols += ["Net_Forecast", "Order_Qty"]
    st.dataframe(
        sheet_df[preview_cols].rename(columns={
            "SubDept": "Sub-Dept", "Sys_Stock": "SOH Now",
            "Pre_Del_Fc": "Fc to Deliv.", "Proj_Stock": "SOH at Del.",
            "Net_Forecast": "Total Forecast", "Order_Qty": "Order Qty"
        }),
        use_container_width=True,
        hide_index=True,
    )

st.divider()

# ── Section 5: Generate ────────────────────────────────────────────────────────
st.subheader("5 — Generate")

_supplier_slug = supplier.replace(" ", "_")
_date_str      = date.today().strftime("%Y%m%d")
fname          = f"SCS_{_supplier_slug}_{_date_str}.xlsx"

excel_bytes = build_excel(sheet_df, c_labels, cycle_dates, order_type)

# Save audit copy to 04_ordering/
output_path = OUTPUT_DIR / fname
OUTPUT_DIR.mkdir(exist_ok=True)
try:
    output_path.write_bytes(excel_bytes)
except PermissionError:
    st.warning(
        f"⚠️ Could not save audit copy to `04_ordering/{fname}` — "
        "the file is open in another program. Close it and click Refresh, "
        "or use the download button below."
    )

st.success(f"✅ Ready — **{fname}**  ({len(sheet_df)} items, {len(cycle_dates)} cycle days)")

if supplier == "Freshlink":
    # ── Standard process: download stock count sheet ───────────────────────────
    st.download_button(
        label="⬇️  Download Stock Count Sheet",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
    st.caption(f"File also saved to: 04_ordering/{fname}")

else:
    # ── Email order format for Bowlsome / Local Kitchen / Simply Tasty ─────────
    # Build order lines — only items with a calculated order qty > 0
    _order_rows = sheet_df[
        sheet_df["Order_Qty"].notna() & (sheet_df["Order_Qty"] > 0)
    ][["Name", "Order_Qty", "_unit"]].copy()

    # Items without SOH get a blank qty (buyer confirms manually)
    _manual_rows = sheet_df[sheet_df["Order_Qty"].isna()][["Name"]].copy()

    _delivery_str = delivery_date.strftime("%A, %d %B %Y")
    _order_str    = date.today().strftime("%A, %d %B %Y")
    _cycle_cover  = " · ".join(d.strftime("%a %d/%m") for d in cycle_dates)

    # Build email body
    _lines = []
    _lines.append(f"Subject: Fruit & Veg Order — {supplier} — Delivery {delivery_date.strftime('%d %b %Y')}")
    _lines.append("")
    _lines.append(f"Hi {supplier} team,")
    _lines.append("")
    _lines.append(
        f"Please see our order for delivery on **{_delivery_str}** "
        f"(covering {_cycle_cover})."
    )
    _lines.append("")
    _lines.append("─" * 52)

    if not _order_rows.empty:
        # Column widths
        _max_name = max(_order_rows["Name"].str.len().max(), 30)
        _lines.append(f"{'Item':<{_max_name}}  Qty")
        _lines.append("─" * (_max_name + 8))
        for _, r in _order_rows.iterrows():
            qty_str = (
                str(int(r["Order_Qty"])) if r.get("_unit") == "whole"
                else f"{r['Order_Qty']:.1f}"
            )
            _lines.append(f"{r['Name']:<{_max_name}}  {qty_str}")
    else:
        _lines.append("(No items with calculated order quantity — check SOH upload.)")

    if not _manual_rows.empty:
        _lines.append("")
        _lines.append("Items requiring manual count (SOH not available):")
        for _, r in _manual_rows.iterrows():
            _lines.append(f"  {r['Name']}")

    _lines.append("")
    _lines.append("─" * 52)
    _lines.append("")
    _lines.append("Thank you,")
    _lines.append("Foodland Wudinna — Fruit & Veg")
    _lines.append(f"Order date: {_order_str}")

    _email_text = "\n".join(_lines)

    st.text_area(
        "📧 Order email — copy and send to supplier",
        value=_email_text,
        height=380,
    )

    st.download_button(
        label="⬇️  Download Stock Count Sheet (audit)",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption(f"Audit file saved to: 04_ordering/{fname}")
