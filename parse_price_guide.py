"""
parse_price_guide.py — Parse the weekly Freshlink price guide Excel
and write supplier_prices_YYYYMMDD.csv for predict.py.

Usage
-----
    python parse_price_guide.py <excel_file>

    Drag-and-drop the Excel onto Launch Parse Price Guide.bat for the same effect.

What it does
------------
1. Reads the 4-column-group layout from the price guide Excel (rows 7+).
2. Extracts (description, qty_string, price) from each group.
3. Skips items where price is missing, zero, or non-numeric (out of season / TBA).
4. Matches each description against price_guide_mapping.csv (exact key lookup).
   Falls back to token-overlap matching against invoice_item_mapping.csv for
   descriptions with 2+ significant tokens.
5. Computes: sell_price = (invoice_price / units_per_invoice) / 0.60
6. Writes 01_data/operational/supplier_prices_YYYYMMDD.csv (Name, sell_price).
7. Prints a match summary — review unmatched items and add them to
   01_data/reference/price_guide_mapping.csv to capture them next time.

Mapping files
-------------
01_data/reference/price_guide_mapping.csv
    Primary lookup: price_guide_key → pos_name + units_per_invoice.
    Edit this file to add new items or correct units.

01_data/reference/invoice_item_mapping.csv
    Fallback: used for token-overlap matching when a description is not in
    price_guide_mapping. May be less accurate for units — check the output.

Output
------
01_data/operational/supplier_prices_YYYYMMDD.csv
    Columns: Name, sell_price
    predict.py picks this up automatically (±3 days of the forecast week start).
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT               = Path(__file__).parent
PG_MAPPING_CSV     = ROOT / "01_data/reference/price_guide_mapping.csv"
INV_MAPPING_CSV    = ROOT / "01_data/reference/invoice_item_mapping.csv"
OUTPUT_DIR         = ROOT / "01_data/operational"

# ── Price guide layout ────────────────────────────────────────────────────────
# Each group: (description_col, qty_col, price_col) — 0-indexed
GROUPS = [
    (0, 2, 3),
    (4, 6, 7),
    (8, 9, 10),
    (11, 12, 13),
]
DATA_START_ROW = 7    # 1-indexed; rows 1-6 are headers / category labels
DATE_ROW       = 4
DATE_COL       = 12   # 0-indexed — cell holds "22.04.26"

GP_TARGET = 0.60      # sell_price = cost_per_unit / GP_TARGET

# ── Token matching stop words ─────────────────────────────────────────────────
_STOP = {
    "the", "a", "an", "and", "or", "of", "for", "in", "on", "to", "at",
    "per", "each", "ea", "x", "s", "p",
    "kg", "g", "ml", "l", "ctn", "pkt", "pk",
    "p/p", "pp", "bch", "lge", "sm", "med", "aust", "usa", "sa",
    "loose", "fresh", "pre", "pack", "packed", "new",
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().upper()


def _tokens(s: str) -> set[str]:
    """
    Extract significant word tokens from a string.

    Filters out: stop words, pure numbers, size tokens (250G, 12KG, 750ML),
    and single-character tokens produced by splitting punctuation like
    apostrophes ("25'S" → "25" + "S").
    """
    parts = re.split(r"[\s\-/,.'\"()\[\]]+", s.upper())
    result = set()
    for p in parts:
        if not p or len(p) < 2:                       # drop single chars
            continue
        if p.lower() in _STOP:                         # case-insensitive stop words
            continue
        if re.fullmatch(r"[\d\.]+", p):                # pure numbers
            continue
        if re.fullmatch(r"\d+\s*[GKML]+", p):          # size tokens: 1KG, 250G, 750ML
            continue
        result.add(p)
    return result


def _norm_qty(qty_str: str) -> str:
    """Normalise a QTY string for use in compound mapping keys. e.g. '15kg ' → '15KG'."""
    return re.sub(r"\s+", "", str(qty_str).strip().upper())


# ── Date parsing ──────────────────────────────────────────────────────────────
def _parse_guide_date(cell_val) -> str:
    """Parse '22.04.26' or '22/04/26' from the guide date cell → 'YYYY-MM-DD'."""
    if cell_val is None:
        return datetime.today().strftime("%Y-%m-%d")
    s = str(cell_val).strip()
    for fmt in ("%d.%m.%y", "%d/%m/%y", "%d-%m-%y", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Last resort: look for the date in a longer string like "DATE: 22.04.26"
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = "20" + y
        try:
            return datetime.strptime(f"{d}.{mo}.{y}", "%d.%m.%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return datetime.today().strftime("%Y-%m-%d")


# ── Excel parsing ─────────────────────────────────────────────────────────────
def read_price_guide(path: Path) -> tuple[str, list[tuple[str, str, float]]]:
    """
    Open the Excel file and return (guide_date_str, items).
    items = list of (description, qty_string, price).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Extract date
    try:
        date_row  = ws[DATE_ROW]
        date_cell = date_row[DATE_COL].value
        guide_date = _parse_guide_date(date_cell)
    except Exception:
        guide_date = datetime.today().strftime("%Y-%m-%d")

    items = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        for desc_col, qty_col, price_col in GROUPS:
            desc  = row[desc_col]  if desc_col  < len(row) else None
            qty   = row[qty_col]   if qty_col   < len(row) else None
            price = row[price_col] if price_col < len(row) else None

            if not desc or not str(desc).strip():
                continue
            desc_str = _norm(str(desc))
            if not desc_str:
                continue

            # Require a numeric price > 0
            if price is None:
                continue
            try:
                price_f = float(price)
            except (ValueError, TypeError):
                continue
            if price_f <= 0:
                continue

            qty_str = str(qty).strip() if qty else ""
            items.append((desc_str, qty_str, price_f))

    return guide_date, items


# ── Mapping loaders ───────────────────────────────────────────────────────────
def _load_pg_mapping() -> dict[str, dict]:
    """Load price_guide_mapping.csv → {normalized_key: {pos_name, units_per_invoice}}."""
    if not PG_MAPPING_CSV.exists():
        return {}
    df = pd.read_csv(PG_MAPPING_CSV)
    df.columns = df.columns.str.strip()
    result = {}
    for _, r in df.iterrows():
        key = _norm(str(r.get("price_guide_key", "")))
        pos = str(r.get("pos_name", "")).strip()
        if key and pos:
            try:
                units = float(r.get("units_per_invoice", 1))
            except (ValueError, TypeError):
                units = 1.0
            result[key] = {"pos_name": pos, "units_per_invoice": units}
    return result


def _load_inv_mapping() -> list[dict]:
    """Load invoice_item_mapping.csv for token-overlap fallback."""
    if not INV_MAPPING_CSV.exists():
        return []
    df = pd.read_csv(INV_MAPPING_CSV)
    df.columns = df.columns.str.strip()
    rows = []
    for _, r in df.iterrows():
        inv_desc = _norm(str(r.get("invoice_description", "")))
        pos_name = str(r.get("pos_name", "")).strip()
        if inv_desc and pos_name:
            try:
                units = float(r.get("units_per_invoice", 1))
            except (ValueError, TypeError):
                units = 1.0
            rows.append({
                "invoice_description": inv_desc,
                "pos_name": pos_name,
                "units_per_invoice": units,
            })
    return rows


# ── Matching ──────────────────────────────────────────────────────────────────
def _token_match(desc: str, inv_mapping: list[dict]) -> dict | None:
    """
    Token-overlap fallback against invoice_item_mapping.
    Requires ≥2 overlapping significant tokens AND ≥50% coverage of desc tokens.
    Only attempted for descriptions with ≥2 significant tokens (single-word
    descriptions are too ambiguous to match safely this way).
    """
    desc_tok = _tokens(desc)
    if len(desc_tok) < 2:
        return None

    best      = None
    best_score = 0.0

    for entry in inv_mapping:
        inv_tok = _tokens(entry["invoice_description"])
        if not inv_tok:
            continue
        overlap   = len(desc_tok & inv_tok)
        coverage  = overlap / len(desc_tok)
        if overlap >= 2 and coverage >= 0.5:
            score = overlap + coverage
            if score > best_score:
                best_score = score
                best = entry

    if best:
        return {"pos_name": best["pos_name"], "units_per_invoice": best["units_per_invoice"]}
    return None


def match_items(
    items: list[tuple[str, str, float]],
    pg_mapping: dict,
    inv_mapping: list[dict],
) -> tuple[list[dict], list[str]]:
    """
    Match price guide items to POS names.

    Lookup order:
    1. Compound key  "DESC QTY"  in price_guide_mapping  (e.g. "KESTRAL 2KG")
    2. Plain key     "DESC"      in price_guide_mapping  (e.g. "KESTRAL")
    3. Token-overlap against invoice_item_mapping (requires ≥2 significant tokens)

    Returns
    -------
    matched   : list of {pos_name, sell_price, source, guide_desc, invoice_price}
    unmatched : list of description strings that could not be matched
    """
    matched   = []
    unmatched = []
    seen_pos  = {}   # pos_name → entry (keep highest-confidence match per POS item)

    for desc, qty_str, price in items:
        # 1. Compound key match (desc + normalised qty)
        compound = f"{desc} {_norm_qty(qty_str)}" if qty_str else None
        mapping  = pg_mapping.get(compound) if compound else None
        source   = "price_guide_mapping (compound key)"

        # 2. Plain key match
        if mapping is None:
            mapping = pg_mapping.get(desc)
            source  = "price_guide_mapping"

        # 3. Token overlap fallback
        if mapping is None:
            mapping = _token_match(desc, inv_mapping)
            source  = "invoice_mapping (token match)"

        if mapping is None:
            unmatched.append(desc)
            continue

        pos_name     = _norm(mapping["pos_name"])
        units        = max(mapping["units_per_invoice"], 0.01)
        cost_per_unit = price / units
        sell_price    = round(cost_per_unit / GP_TARGET, 4)

        entry = {
            "pos_name":     pos_name,
            "sell_price":   sell_price,
            "source":       source,
            "guide_desc":   desc,
            "invoice_price": price,
            "units":        units,
            "cost_per_unit": round(cost_per_unit, 4),
        }

        # If we already matched a different description to the same POS item,
        # keep both (the CSV will have duplicates for the same Name — predict.py
        # takes the first, so we deduplicate here by keeping the first match).
        if pos_name not in seen_pos:
            seen_pos[pos_name] = entry
            matched.append(entry)
        # else: silently skip duplicate POS item matches

    return matched, unmatched


# ── Output ────────────────────────────────────────────────────────────────────
def write_output(matched: list[dict], guide_date: str) -> Path:
    """Write supplier_prices_YYYYMMDD.csv and return its path."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    date_stamp = guide_date.replace("-", "")
    out_path   = OUTPUT_DIR / f"supplier_prices_{date_stamp}.csv"

    rows = [{"Name": e["pos_name"], "sell_price": e["sell_price"]} for e in matched]
    df   = pd.DataFrame(rows, columns=["Name", "sell_price"])
    df.to_csv(out_path, index=False)
    return out_path


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Parse Freshlink price guide → supplier_prices_YYYYMMDD.csv"
    )
    parser.add_argument("excel_file", help="Path to the Freshlink price guide Excel (.xlsx)")
    parser.add_argument(
        "--gp", type=float, default=GP_TARGET,
        help=f"GP target as a decimal (default: {GP_TARGET} = 40%% GP)",
    )
    args = parser.parse_args()

    xl_path = Path(args.excel_file)
    if not xl_path.exists():
        print(f"ERROR: File not found: {xl_path}", file=sys.stderr)
        sys.exit(1)

    gp_target = args.gp

    print(f"Parsing: {xl_path.name}")
    print()

    # Parse Excel
    try:
        guide_date, items = read_price_guide(xl_path)
    except Exception as e:
        print(f"ERROR reading price guide: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"  Guide date:       {guide_date}")
    print(f"  Items in guide:   {len(items)}")

    # Load mapping files
    pg_mapping  = _load_pg_mapping()
    inv_mapping = _load_inv_mapping()

    # Match items
    matched, unmatched = match_items(items, pg_mapping, inv_mapping)

    print(f"  Matched:          {len(matched)}")
    print(f"  Unmatched:        {len(unmatched)}")
    print()

    if not matched:
        print("WARNING: No items matched. Check that price_guide_mapping.csv exists and is populated.")
        sys.exit(1)

    # Write output
    out_path = write_output(matched, guide_date)
    print(f"  Output written:   {out_path.relative_to(ROOT)}")
    print()

    # ── Match detail ─────────────────────────────────────────────────────────
    col_w = max(len(e["guide_desc"]) for e in matched)
    print(f"  {'Description':<{col_w}}  {'POS Name':<45}  {'Units':>5}  {'Cost/U':>7}  {'Sell':>7}  Source")
    print(f"  {'-'*col_w}  {'-'*45}  {'-'*5}  {'-'*7}  {'-'*7}  ------")
    for e in sorted(matched, key=lambda x: x["guide_desc"]):
        print(
            f"  {e['guide_desc']:<{col_w}}  {e['pos_name']:<45}  "
            f"{e['units']:>5.1f}  {e['cost_per_unit']:>7.2f}  {e['sell_price']:>7.2f}  "
            f"{'PG' if 'price_guide' in e['source'] else 'INV'}"
        )

    # ── Unmatched ─────────────────────────────────────────────────────────────
    if unmatched:
        print()
        print(f"  Unmatched descriptions ({len(unmatched)}):")
        print(f"  Add these to 01_data/reference/price_guide_mapping.csv to capture them next run.")
        print()
        for d in sorted(unmatched):
            print(f"    {d}")

    print()
    print("Done.")


if __name__ == "__main__":
    main()
